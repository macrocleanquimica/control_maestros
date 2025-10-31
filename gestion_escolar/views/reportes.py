import json
import openpyxl
from datetime import date

from django.shortcuts import render
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.db.models import Q, Count
from django.db.models.functions import Upper, Trim

from ..models import Maestro, Zona, Escuela

@permission_required('gestion_escolar.acceder_reportes', raise_exception=True)
def reportes_dashboard(request):
    context = {
        'titulo': 'Dashboard de Reportes'
    }
    return render(request, 'gestion_escolar/reportes_dashboard.html', context)

@login_required
def reporte_personal_fuera_adscripcion(request):
    personal_qs = Maestro.objects.annotate(
        techo_f_clean=Trim(Upper('techo_f')),
        id_escuela_clean=Trim(Upper('id_escuela__id_escuela'))
    ).exclude(techo_f__isnull=True).exclude(techo_f='')

    personal_fuera_adscripcion = [p for p in personal_qs if p.techo_f_clean != p.id_escuela_clean]

    context = {
        'personal_fuera_adscripcion': personal_fuera_adscripcion,
        'titulo': 'Reporte de Personal Fuera de Adscripción'
    }
    return render(request, 'gestion_escolar/reporte_fuera_adscripcion.html', context)

@login_required
def export_maestro_excel(request, pk):
    maestro = get_object_or_404(Maestro, id_maestro=pk)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Detalle_{maestro.id_maestro}"

    data = {
        "ID Maestro": maestro.id_maestro,
        "Nombre Completo": f'{maestro.nombres} {maestro.a_paterno} {maestro.a_materno}',
        "CURP": maestro.curp,
        "RFC": maestro.rfc,
        "Sexo": maestro.get_sexo_display(),
        "Estado Civil": maestro.get_est_civil_display(),
        "Fecha de Nacimiento": maestro.fecha_nacimiento,
        "Techo Financiero": maestro.techo_f,
        "C.C.T.": maestro.id_escuela.id_escuela if maestro.id_escuela else 'N/A',
        "Nombre del Centro de Trabajo": maestro.id_escuela.nombre_ct if maestro.id_escuela else 'N/A',
        "Zona Escolar": maestro.id_escuela.zona_esc.numero if maestro.id_escuela and maestro.id_escuela.zona_esc else 'N/A',
        "Clave Presupuestal": maestro.clave_presupuestal,
        "Categoría": str(maestro.categog) if maestro.categog else '',
        "Código": maestro.codigo,
        "Fecha de Ingreso": maestro.fecha_ingreso,
        "Fecha de Promoción": maestro.fecha_promocion,
        "Formación Académica": maestro.form_academica,
        "Horario": maestro.horario,
        "Función": maestro.get_funcion_display(),
        "Nivel de Estudio": maestro.get_nivel_estudio_display(),
        "Domicilio Particular": maestro.domicilio_part,
        "Población": maestro.poblacion,
        "Código Postal": maestro.codigo_postal,
        "Teléfono": maestro.telefono,
        "Email": maestro.email,
        "Status": maestro.get_status_display(),
        "Observaciones": maestro.observaciones,
    }

    for key, value in data.items():
        if isinstance(value, date):
            value = value.strftime("%d/%m/%Y")
        ws.append([key, value])

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="detalle_{maestro.a_paterno}_{maestro.id_maestro}.xlsx"'},
    )
    wb.save(response)

    return response

@login_required
def exportar_maestros_excel(request):
    filtro = request.GET.get('filtro', '')
    funcion = request.GET.get('funcion', '')

    maestros_qs = Maestro.objects.select_related('id_escuela', 'id_escuela__zona_esc', 'categog').all().order_by('a_paterno', 'a_materno', 'nombres')

    if funcion:
        # This mapping should ideally be in a more centralized place
        funcion_mapping = {
            'DIRECTOR': {'values': ['DIRECTOR', 'DIRECTOR (A)']},
            # ... add all other mappings here ...
        }
        funcion_info = funcion_mapping.get(funcion)
        if funcion_info:
            maestros_qs = maestros_qs.filter(funcion__in=funcion_info['values'])

    if filtro:
        maestros_qs = maestros_qs.filter(
            Q(nombres__icontains=filtro) |
            Q(a_paterno__icontains=filtro) |
            Q(a_materno__icontains=filtro) |
            Q(rfc__icontains=filtro) |
            Q(curp__icontains=filtro) |
            Q(id_maestro__icontains=filtro) |
            Q(id_escuela__nombre_ct__icontains=filtro) |
            Q(id_escuela__id_escuela__icontains=filtro) |
            Q(categog__descripcion__icontains=filtro)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maestros"

    headers = [
        "ID Maestro", "Nombre(s)", "Apellido Paterno", "Apellido Materno", "RFC", "CURP",
        "Sexo", "Estado Civil", "Fecha Nacimiento", "Techo Financiero",
        "CCT", "Nombre del CT", "Zona Escolar", "Función", "Categoría",
        "Clave Presupuestal", "Código", "Fecha Ingreso", "Fecha Promoción",
        "Formación Académica", "Horario", "Nivel de Estudio", "Domicilio Particular",
        "Población", "Código Postal", "Teléfono", "Email", "Status", "Observaciones"
    ]
    ws.append(headers)

    for maestro in maestros_qs:
        escuela = maestro.id_escuela
        zona_numero = ''
        if escuela and escuela.zona_esc:
            zona_numero = escuela.zona_esc.numero

        row = [
            maestro.id_maestro, maestro.nombres, maestro.a_paterno, maestro.a_materno, maestro.rfc, maestro.curp,
            maestro.get_sexo_display(), maestro.get_est_civil_display(),
            maestro.fecha_nacimiento.strftime("%Y-%m-%d") if maestro.fecha_nacimiento else '',
            maestro.techo_f, escuela.id_escuela if escuela else '', escuela.nombre_ct if escuela else '', zona_numero,
            maestro.get_funcion_display(), maestro.categog.descripcion if maestro.categog else '',
            maestro.clave_presupuestal, maestro.codigo,
            maestro.fecha_ingreso.strftime("%Y-%m-%d") if maestro.fecha_ingreso else '',
            maestro.fecha_promocion.strftime("%Y-%m-%d") if maestro.fecha_promocion else '',
            maestro.form_academica, maestro.horario, maestro.get_nivel_estudio_display(),
            maestro.domicilio_part, maestro.poblacion, maestro.codigo_postal, maestro.telefono, maestro.email,
            maestro.get_status_display(), maestro.observaciones,
        ]
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="reporte_maestros.xlsx"'},
    )
    wb.save(response)

    return response

@login_required
def reporte_distribucion_funcion(request):
    maestros_qs = Maestro.objects.all()

    zona_id = request.GET.get('zona')
    escuela_id = request.GET.get('escuela')

    if zona_id:
        maestros_qs = maestros_qs.filter(id_escuela__zona_esc_id=zona_id)
    
    if escuela_id:
        maestros_qs = maestros_qs.filter(id_escuela_id=escuela_id)

    distribucion = maestros_qs.values('funcion').annotate(total=Count('funcion')).order_by('-total')

    labels = [d['funcion'] for d in distribucion]
    data = [d['total'] for d in distribucion]

    zonas = Zona.objects.all()
    escuelas = Escuela.objects.all()

    context = {
        'titulo': 'Distribución de Personal por Función',
        'labels': json.dumps(labels),
        'data': json.dumps(data),
        'zonas': zonas,
        'escuelas': escuelas,
        'selected_zona': int(zona_id) if zona_id else None,
        'selected_escuela': int(escuela_id) if escuela_id else None,
    }
    return render(request, 'gestion_escolar/reporte_distribucion_funcion.html', context)

@login_required
def export_personal_fuera_adscripcion_excel(request):
    """Exporta el reporte de personal fuera de adscripción a un archivo Excel, aplicando un filtro de búsqueda."""
    filtro = request.GET.get('filtro', '')

    # 1. Obtener el queryset base
    personal_qs = Maestro.objects.annotate(
        techo_f_clean=Trim(Upper('techo_f')),
        id_escuela_clean=Trim(Upper('id_escuela__id_escuela'))
    ).exclude(techo_f__isnull=True).exclude(techo_f='').select_related('id_escuela', 'id_escuela__zona_esc', 'categog')

    personal_fuera_adscripcion = [p for p in personal_qs if p.techo_f_clean != p.id_escuela_clean]
    
    # Convertir la lista de objetos a un queryset para poder filtrar más
    pks = [p.pk for p in personal_fuera_adscripcion]
    queryset = Maestro.objects.filter(pk__in=pks)

    # 2. Aplicar el filtro si existe
    if filtro:
        queryset = queryset.filter(
            Q(nombres__icontains=filtro) |
            Q(a_paterno__icontains=filtro) |
            Q(a_materno__icontains=filtro) |
            Q(clave_presupuestal__icontains=filtro) |
            Q(id_escuela__id_escuela__icontains=filtro) | # CCT Físico
            Q(techo_f__icontains=filtro) # CCT de Pago
        )

    # 3. Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personal Fuera de Adscripción"

    # 4. Definir los encabezados
    headers = [
        "ID Maestro", "Nombre(s)", "Apellido Paterno", "Apellido Materno", "RFC", "CURP",
        "Sexo", "Estado Civil", "Fecha Nacimiento", "Techo Financiero",
        "CCT Físico", "Nombre del CT Físico", "Zona Escolar", "Función", "Categoría",
        "Clave Presupuestal", "Código", "Fecha Ingreso", "Fecha Promoción",
        "Formación Académica", "Horario", "Nivel de Estudio", "Domicilio Particular",
        "Población", "Código Postal", "Teléfono", "Email", "Status", "Observaciones"
    ]
    ws.append(headers)

    # 5. Escribir los datos de cada maestro
    for maestro in queryset:
        escuela = maestro.id_escuela
        zona_numero = ''
        if escuela and escuela.zona_esc:
            zona_numero = escuela.zona_esc.numero

        row = [
            maestro.id_maestro, maestro.nombres, maestro.a_paterno, maestro.a_materno, maestro.rfc, maestro.curp,
            maestro.get_sexo_display(), maestro.get_est_civil_display(),
            maestro.fecha_nacimiento.strftime("%Y-%m-%d") if maestro.fecha_nacimiento else '',
            maestro.techo_f, 
            escuela.id_escuela if escuela else '', 
            escuela.nombre_ct if escuela else '', 
            zona_numero,
            maestro.get_funcion_display(), 
            maestro.categog.descripcion if maestro.categog else '',
            maestro.clave_presupuestal, maestro.codigo,
            maestro.fecha_ingreso.strftime("%Y-%m-%d") if maestro.fecha_ingreso else '',
            maestro.fecha_promocion.strftime("%Y-%m-%d") if maestro.fecha_promocion else '',
            maestro.form_academica, maestro.horario, maestro.get_nivel_estudio_display(),
            maestro.domicilio_part, maestro.poblacion, maestro.codigo_postal, maestro.telefono, maestro.email,
            maestro.get_status_display(), maestro.observaciones,
        ]
        ws.append(row)

    # 6. Preparar la respuesta para la descarga
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="reporte_personal_fuera_adscripcion.xlsx"'},
    )
    wb.save(response)

    return response

