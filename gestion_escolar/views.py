from django.core.exceptions import PermissionDenied
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm
from django.contrib.auth import update_session_auth_hash

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin, PermissionRequiredMixin
from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.utils import timezone

from django.contrib.auth.models import Group, User, Permission
from django.contrib.contenttypes.models import ContentType
from .forms import (ZonaForm, EscuelaForm, MaestroForm, CategoriaForm, TramiteForm, 
                   SignUpForm, VacanciaForm, DocumentoExpedienteForm, 
                   CustomUserChangeForm, AsignarDirectorForm, PendienteForm, CorrespondenciaForm, RegistroCorrespondenciaForm,
                   RolePermissionForm)
from .models import (Zona, Escuela, Maestro, Categoria, MotivoTramite, 
                   PlantillaTramite, Prelacion, LoteReporteVacancia, Vacancia, 
                   TipoApreciacion, Historial, DocumentoExpediente, 
                   Correspondencia, Notificacion, Pendiente, RegistroCorrespondencia, KardexMovimiento)
from django.db import transaction
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
import csv
import json
import os
import openpyxl
from docxtpl import DocxTemplate
from datetime import datetime, date
from django.conf import settings
import gspread
from google.oauth2 import service_account


# Helper function to get full name
def get_full_name(maestro):
    if not maestro: return ""
    return f"{maestro.nombres or ''} {maestro.a_paterno or ''} {maestro.a_materno or ''}".strip()

# Helper function to get school info
def get_school_info(escuela):
    if not escuela: return {'nombre_ct': '', 'id_escuela': '', 'turno': '', 'domicilio': '', 'zona_economica': '', 'zona_esc_numero': '', 'region': '', 'u_d': '', 'sostenimiento': ''}
    return {
        'nombre_ct': escuela.nombre_ct or '',
        'id_escuela': escuela.id_escuela or '',
        'turno': escuela.get_turno_display() or '',
        'domicilio': escuela.domicilio or '',
        'zona_economica': escuela.zona_economica or '',
        'zona_esc_numero': escuela.zona_esc.numero if escuela.zona_esc else '',
        'region': escuela.region or '',
        'u_d': escuela.u_d or '',
        'sostenimiento': escuela.get_sostenimiento_display() or '',
    }

# Helper function to get director
def get_director_info(escuela):
    if not escuela: return {'nombre': 'DIRECTOR NO ENCONTRADO', 'nivel': ''}
    director = Maestro.objects.filter(id_escuela=escuela, funcion__in=['DIRECTOR', 'DIRECTOR (A)']).first()
    if director:
        return {'nombre': get_full_name(director), 'nivel': director.nivel_estudio or ''}
    return {'nombre': 'DIRECTOR NO ENCONTRADO', 'nivel': ''}

# Helper function to get supervisor
def get_supervisor_info(zona):
    if not zona: return {'nombre': 'SUPERVISOR NO ENCONTRADO', 'nivel': ''}
    supervisor = Maestro.objects.filter(id_escuela__zona_esc=zona, funcion__in=['SUPERVISOR', 'SUPERVISOR (A)', 'SUPERVISOR(A)']).first()
    if supervisor:
        return {'nombre': get_full_name(supervisor), 'nivel': supervisor.nivel_estudio or ''}
    return {'nombre': 'SUPERVISOR NO ENCONTRADO', 'nivel': ''}

# Helper function to get user initials
def get_user_initials(user):
    if not user:
        return ""
    
    # Intenta obtener el nombre completo del perfil del usuario
    full_name = user.get_full_name()
    
    # Si no hay nombre completo, intenta con first_name y last_name
    if not full_name:
        full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()

    if not full_name:
        return user.username[0].lower() if user.username else ''

    parts = full_name.split()
    initials = "".join([part[0] for part in parts if part])
    return initials.lower()


# Helper function to calculate month difference similar to VBA DateDiff("m", ...)
def get_month_diff(d1, d2):
    # Ensure d1 is earlier than d2
    if d1 > d2:
        d1, d2 = d2, d1

    months = (d2.year - d1.year) * 12 + d2.month - d1.month
    if d2.day < d1.day:
        months -= 1
    return months

def numero_a_letras_general(num):
    unidades = ["", "uno", "dos", "tres", "cuatro", "cinco", "seis", "siete", "ocho", "nueve"]
    dieces = ["", "diez", "veinte", "treinta", "cuarenta", "cincuenta", "sesenta", "setenta", "ochenta", "noventa"]
    centenas = ["", "ciento", "doscientos", "trescientos", "cuatrocientos", "quinientos", "seiscientos", "setecientos", "ochocientos", "novecientos"]
    especiales = {
        11: "once", 12: "doce", 13: "trece", 14: "catorce", 15: "quince",
        16: "dieciséis", 17: "diecisiete", 18: "dieciocho", 19: "diecinueve",
        21: "veintiuno", 22: "veintidós", 23: "veintitrés", 24: "veinticuatro",
        25: "veinticinco", 26: "veintiséis", 27: "veintisiete", 28: "veintiocho", 29: "veintinueve"
    }

    if num == 0: return "cero"
    if num in especiales: return especiales[num]

    if num < 10: return unidades[num]
    if num < 30:
        if num % 10 == 0: return dieces[num // 10]
        return dieces[num // 10] + " y " + unidades[num % 10]
    if num < 100:
        if num % 10 == 0: return dieces[num // 10]
        return dieces[num // 10] + " y " + unidades[num % 10]
    if num == 100: return "cien"
    if num < 1000:
        if num % 100 == 0: return centenas[num // 100]
        return centenas[num // 100] + " " + numero_a_letras_general(num % 100)
    if num == 1000: return "mil"
    if num < 2000:
        return "mil " + numero_a_letras_general(num - 1000)
    if num < 1000000: # For years like 2025
        miles = num // 1000
        resto = num % 1000
        if miles == 1:
            letras = "mil"
        else:
            letras = numero_a_letras_general(miles) + " mil"
        if resto > 0:
            letras += " " + numero_a_letras_general(resto)
        return letras
    return str(num) # Fallback for very large numbers

def convertir_fecha_a_letras(fecha):
    dia = numero_a_letras_general(fecha.day)
    meses_letras = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    mes = meses_letras[fecha.month - 1]
    anio = numero_a_letras_general(fecha.year)

    return f"{dia} de {mes} del {anio}"

def format_date_for_solicitud_asignacion(fecha):
    if not fecha:
        return ''
    meses_espanol = [
        "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
        "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
    ]
    dia = fecha.day
    mes = meses_espanol[fecha.month - 1]
    anio = fecha.year
    return f"{dia:02d} DE {mes} DEL {anio}"

def serialize_form_data(cleaned_data):
    """
    Converts form.cleaned_data into a JSON-serializable dictionary,
    handling model instances and date objects.
    """
    serialized_data = {}
    for key, value in cleaned_data.items():
        if hasattr(value, 'pk'):  # It's a Django model instance
            serialized_data[key] = value.pk
        elif isinstance(value, (date, datetime)):
            serialized_data[key] = value.isoformat()
        elif isinstance(value, dict): # Recursively handle nested dictionaries
            serialized_data[key] = serialize_form_data(value)
        elif isinstance(value, list): # Handle lists of serializable items
            serialized_data[key] = [
                item.pk if hasattr(item, 'pk') else
                item.isoformat() if isinstance(item, (date, datetime)) else
                item
                for item in value
            ]
        else:
            serialized_data[key] = value
    return serialized_data

# Helper function to send data to Google Sheet - VERSIÓN CORREGIDA
def send_to_google_sheet(row_data):
    """
    Función corregida para enviar datos a Google Sheets
    """
    print("DEBUG GS: Iniciando envío a Google Sheets...")
    
    try:
        # Verificar configuraciones
        if not hasattr(settings, 'GOOGLE_SHEETS_CREDENTIALS'):
            return False, "Configuración de credenciales no encontrada"
        
        creds_json = settings.GOOGLE_SHEETS_CREDENTIALS
        
        # Verificar campos esenciales
        if not creds_json.get('private_key') or not creds_json.get('client_email'):
            return False, "Credenciales incompletas - falta private_key o client_email"
        
        print(f"DEBUG GS: Usando cuenta: {creds_json['client_email']}")
        
        # Configurar scopes
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        
        # Crear credenciales
        credentials = service_account.Credentials.from_service_account_info(
            creds_json,
            scopes=SCOPES
        )
        
        # Autorizar
        gc = gspread.authorize(credentials)
        print("DEBUG GS: ✅ Autenticación exitosa")
        
        # Abrir la hoja
        spreadsheet = gc.open_by_key(settings.GOOGLE_SHEET_ID)  # open_by_key en lugar de open_by_id
        worksheet = spreadsheet.worksheet(settings.GOOGLE_SHEET_WORKSHEET_NAME)
        print("DEBUG GS: ✅ Hoja abierta correctamente")
        
        # Preparar datos
        cleaned_data = []
        for item in row_data:
            if item is None:
                cleaned_data.append('')
            elif isinstance(item, (date, datetime)):
                cleaned_data.append(item.strftime('%Y-%m-%d'))
            else:
                cleaned_data.append(str(item))
        
        # Enviar datos
        worksheet.append_row(cleaned_data)
        print(f"DEBUG GS: ✅ Fila agregada: {cleaned_data}")
        
        return True, "Datos enviados correctamente a Google Sheets"
        
    except gspread.exceptions.SpreadsheetNotFound:
        error_msg = "Google Sheet no encontrado. Verifica el GOOGLE_SHEET_ID."
        print(f"DEBUG GS: ❌ {error_msg}")
        return False, error_msg
        
    except gspread.exceptions.WorksheetNotFound:
        error_msg = f"Hoja '{settings.GOOGLE_SHEET_WORKSHEET_NAME}' no encontrada."
        print(f"DEBUG GS: ❌ {error_msg}")
        return False, error_msg
        
    except Exception as e:
        error_msg = f"Error: {str(e)}"
        print(f"DEBUG GS: ❌ {error_msg}")
        return False, error_msg

# Main Word generation function
def generate_word_document(form_data, plantilla_tramite, user):
    try:
        # --- Lógica para seleccionar plantilla especial para desubicados ---
        maestro_titular = form_data.get('maestro_titular')
        template_name_upper = plantilla_tramite.nombre.upper().strip()
        ruta_plantilla_final = plantilla_tramite.ruta_archivo

        is_desubicado = False
        if maestro_titular and maestro_titular.techo_f and maestro_titular.id_escuela:
            if maestro_titular.techo_f.strip().upper() != maestro_titular.id_escuela.id_escuela.strip().upper():
                is_desubicado = True

        # Mapeo de plantillas normales a sus versiones para desubicados
        plantillas_especiales = {
            "REINGRESO": "REINGRESODESUBICADO.docx",
            "FILIACION": "FILIACIONDESUBICADO.docx",
        }

        # Si la plantilla actual tiene una versión especial y el maestro está desubicado, la usamos
        if template_name_upper in plantillas_especiales and is_desubicado:
            nueva_plantilla = plantillas_especiales[template_name_upper]
            ruta_plantilla_final = nueva_plantilla
            print(f"DEBUG: Maestro desubicado detectado para {template_name_upper}. Usando plantilla especial: {ruta_plantilla_final}")

        # --- Construct absolute path for the template ---
        template_path = os.path.join(settings.BASE_DIR, 'tramites', 'Plantillas', 'Word', ruta_plantilla_final)
        doc = DocxTemplate(template_path)

        # --- Gather Data ---
        maestro_interino = form_data.get('maestro_interino')
        motivo_tramite_obj = form_data.get('motivo_tramite')

        # Data from Maestro Titular
        nombre_titular = get_full_name(maestro_titular)
        curp_titular = maestro_titular.curp or '' if maestro_titular else ''
        rfc_titular = maestro_titular.rfc or '' if maestro_titular else ''
        categoria_titular = maestro_titular.categog.descripcion if maestro_titular and maestro_titular.categog else ''
        presupuestal_titular = maestro_titular.clave_presupuestal or '' if maestro_titular else ''
        techo_financiero_titular = maestro_titular.techo_f or '' if maestro_titular else ''
        funcion_titular = maestro_titular.funcion or '' if maestro_titular else ''

        # Data from Maestro Interino
        nombre_interino = get_full_name(maestro_interino)
        curp_interino = maestro_interino.curp or '' if maestro_interino else ''
        rfc_interino = maestro_interino.rfc or '' if maestro_interino else ''
        domicilio_part_interino = maestro_interino.domicilio_part or '' if maestro_interino else ''
        codigo_postal_interino = maestro_interino.codigo_postal or '' if maestro_interino else ''
        poblacion_interino = maestro_interino.poblacion or '' if maestro_interino else ''
        telefono_interino = maestro_interino.telefono or '' if maestro_interino else ''
        codigo_interino = maestro_interino.codigo or '' if maestro_interino else ''
        paterno_interino = maestro_interino.a_paterno or '' if maestro_interino else ''
        materno_interino = maestro_interino.a_materno or '' if maestro_interino else ''
        nombre_interino_solo = maestro_interino.nombres or '' if maestro_interino else ''
        formacion_academica_interino = maestro_interino.form_academica or '' if maestro_interino else ''
        
        # Lógica para presupuestal_interino
        presupuestal_interino = presupuestal_titular
        if motivo_tramite_obj and presupuestal_titular and len(presupuestal_titular) >= 2:
            motivo_text = motivo_tramite_obj.motivo_tramite.upper().strip()
            if motivo_text == "BECA COMISIÓN" or motivo_text == "PRORROGA DE BECA COMISION":
                presupuestal_interino = "48" + presupuestal_titular[2:]
            elif motivo_text == "LIC. DE GRAVIDEZ":
                presupuestal_interino = "24" + presupuestal_titular[2:]
            elif motivo_text == "LIC. PREPENSIONARIA":
                presupuestal_interino = "15" + presupuestal_titular[2:]
            elif motivo_text == "PREJUBILATORIO":
                presupuestal_interino = "15" + presupuestal_titular[2:]
        
        # Lógica para funcion_interino
        funcion_interino = maestro_titular.funcion or '' if maestro_titular else ''

        # Data from Form fields
        folio = form_data.get('folio') or ''
        fecha_efecto1 = form_data.get('fecha_efecto1')
        fecha_efecto2 = form_data.get('fecha_efecto2')
        fecha_efecto3 = form_data.get('fecha_efecto3')
        fecha_efecto4 = form_data.get('fecha_efecto4')
        motivo_movimiento = motivo_tramite_obj.motivo_tramite if motivo_tramite_obj else ''
        observaciones = form_data.get('observaciones') or ''
        quincena_inicial = form_data.get('quincena_inicial') or ''
        quincena_final = form_data.get('quincena_final') or ''

        # Lógica para calcular tipo_movimiento_interino
        motivo_tramite_text = motivo_tramite_obj.motivo_tramite.upper().strip() if motivo_tramite_obj else ''
        tipo_movimiento_interino = ""

        if motivo_tramite_text == "LIC. DE GRAVIDEZ":
            tipo_movimiento_interino = "ALTA INTERINA EN GRAVIDEZ"
        elif motivo_tramite_text == "LIC. POR PASAR A OTRO EMPLEO":
            tipo_movimiento_interino = "ALTA INICIAL POR PROMOCIÓN O ADMISIÓN"
        else:
            if not fecha_efecto3 or not fecha_efecto4:
                tipo_movimiento_interino = "FECHAS INSUFICIENTES"
            else:
                diferencia_meses = get_month_diff(fecha_efecto3, fecha_efecto4)

                if motivo_tramite_text in ["LIC. PREPENSIONARIA", "PREJUBILATORIO"]:
                    if diferencia_meses < 6:
                        tipo_movimiento_interino = "ALTA EN PENSION"
                    else:
                        tipo_movimiento_interino = "ALTA PROVISIONAL"
                elif motivo_tramite_text in ["BECA COMISIÓN", "PRORROGA DE BECA COMISION", "PRÓRROGA DE BECA COMISIÓN"]:
                    if diferencia_meses < 6:
                        tipo_movimiento_interino = "SUSTITUTO BECARIO"
                    else:
                        tipo_movimiento_interino = "ALTA PROVISIONAL"
                elif motivo_tramite_text in ["BAJA POR DEFUNCIÓN", "LIC. POR ASUNTOS PARTICULARES", "LIC. POR COM. SINDICAL", "PRORROGA DE LIC. POR COM. SINDICAL"]:
                    if diferencia_meses < 6:
                        tipo_movimiento_interino = "ALTA INTERINA LIMITADA"
                    else:
                        tipo_movimiento_interino = "ALTA PROVISIONAL"
                elif motivo_tramite_text == "JUBILACIÓN":
                    if diferencia_meses < 6:
                        tipo_movimiento_interino = "ALTA INTERINA LIMITADA EN VACANTE DEFINITIVA"
                    else:
                        tipo_movimiento_interino = "ALTA PROVISIONAL EN VACante DEFINITIVA"
                else:
                    tipo_movimiento_interino = "NO PROCEDENTE"

        # Dynamic Dates
        today = datetime.now()
        meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
        f_hoy = f"{today.day} de {meses[today.month - 1]} del {today.year}"
        f_hoy_letras = convertir_fecha_a_letras(today)

        # --- OBTENER DATOS DE AMBAS ESCUELAS: ADSCRIPCIÓN Y PAGO ---
        # 1. Escuela de Adscripción (donde trabaja físicamente)
        escuela_adscripcion = None
        if maestro_titular:
            escuela_adscripcion = maestro_titular.id_escuela
        
        escuela_adscripcion_info = get_school_info(escuela_adscripcion)

        # Corrección: Asegurarse de que las variables de director y supervisor siempre existan
        if escuela_adscripcion:
            director_adscripcion_info = get_director_info(escuela_adscripcion)
            supervisor_adscripcion_info = get_supervisor_info(escuela_adscripcion.zona_esc)
        else:
            director_adscripcion_info = {'nombre': 'DIRECTOR NO ENCONTRADO', 'nivel': ''}
            supervisor_adscripcion_info = {'nombre': 'SUPERVISOR NO ENCONTRADO', 'nivel': ''}
        # 2. Escuela de Pago (donde cobra, según Techo Financiero)
        escuela_pago = None
        if maestro_titular and maestro_titular.techo_f:
            try:
                escuela_pago = Escuela.objects.get(id_escuela=maestro_titular.techo_f)
            except Escuela.DoesNotExist:
                # Si el techo_f no corresponde a una escuela registrada, se deja como None
                escuela_pago = None
        
        escuela_pago_info = get_school_info(escuela_pago)

        # 3. Autoridades de la Escuela de Pago (Techo Financiero)
        if escuela_pago:
            director_pago_info = get_director_info(escuela_pago)
            supervisor_pago_info = get_supervisor_info(escuela_pago.zona_esc)
        else:
            # Si no hay escuela de pago, usar valores por defecto
            director_pago_info = {'nombre': 'DIRECTOR (PAGO) NO ENCONTRADO', 'nivel': ''}
            supervisor_pago_info = {'nombre': 'SUPERVISOR (PAGO) NO ENCONTRADO', 'nivel': ''}

        quincena_inicial = form_data.get('quincena_inicial') or ''
        quincena_final = form_data.get('quincena_final') or ''

        # Extract day, month, year for [I_Dia], [I_Mes], [I_Ano]
        i_dia = f"{fecha_efecto3.day:02d}" if fecha_efecto3 else ''
        i_mes = f"{fecha_efecto3.month:02d}" if fecha_efecto3 else '' # Month as number
        i_ano = fecha_efecto3.year if fecha_efecto3 else ''

        # Extract day, month, year for [F_Dia], [F_Mes], [F_Ano]
        f_dia = f"{fecha_efecto4.day:02d}" if fecha_efecto4 else ''
        f_mes = f"{fecha_efecto4.month:02d}" if fecha_efecto4 else '' # Month as number
        f_ano = fecha_efecto4.year if fecha_efecto4 else ''

        # Obtener datos de prelación del formulario
        no_prel = form_data.get('no_prel_display') or ''
        folio_prel = form_data.get('folio_prel_display') or ''
        tipo_val = form_data.get('tipo_val_display') or ''

        # --- Generate initials ---
        quienlohizo = get_user_initials(user)

        # --- Context for docxtpl ---
        context = {
            'quienlohizo': quienlohizo,
            'Nombre_Titular': nombre_titular,
            'CURP_Titular': curp_titular,
            'RFC_Titular': rfc_titular,
            'Categoria_Titular': categoria_titular,
            'Presupuestal_Titular': presupuestal_titular,
            'Techo_Financiero': techo_financiero_titular,
            'Funcion_Titular': funcion_titular,
            
            # --- Variables para la Escuela de Adscripción (donde trabaja) ---
            'Clave_CT': escuela_adscripcion_info['id_escuela'],
            'Nombre_CT': escuela_adscripcion_info['nombre_ct'],
            'Turno': escuela_adscripcion_info['turno'],
            'Domicilio_CT': escuela_adscripcion_info['domicilio'],
            'Z_economica': escuela_adscripcion_info['zona_economica'],
            'Z_Escolar': escuela_adscripcion_info['zona_esc_numero'],
            'Poblacion': escuela_adscripcion_info['region'],
            'U_D': escuela_adscripcion_info['u_d'],
            'Sostenimiento': escuela_adscripcion_info['sostenimiento'],
            'Nom_CTCompleto': escuela_adscripcion_info['nombre_ct'],

            # --- NUEVAS Variables para la Escuela de Pago (Techo Financiero) ---
            'Clave_CT_Techo_F': escuela_pago_info['id_escuela'],
            'Nombre_CT_Techo_F': escuela_pago_info['nombre_ct'],
            'Turno_Techo_F': escuela_pago_info['turno'],
            'Domicilio_CT_Techo_F': escuela_pago_info['domicilio'],
            'Poblacion_Techo_F': escuela_pago_info['region'],
            'Nom_CT_Techo_F_Completo': escuela_pago_info['nombre_ct'], # <-- VARIABLE AÑADIDA

            'T_Movimiento': motivo_movimiento,
            'Efecto_1': fecha_efecto1.strftime("%d/%m/%Y") if fecha_efecto1 else '',
            'Efecto_2': fecha_efecto2.strftime("%d/%m/%Y") if fecha_efecto2 else '',
            'Efecto_3': format_date_for_solicitud_asignacion(fecha_efecto3) if plantilla_tramite.nombre == "SOLICITUD DE ASIGNACION" else (fecha_efecto3.strftime("%d/%m/%Y") if fecha_efecto3 else ''),
            'Efecto_4': format_date_for_solicitud_asignacion(fecha_efecto4) if plantilla_tramite.nombre == "SOLICITUD DE ASIGNACION" else (fecha_efecto4.strftime("%d/%m/%Y") if fecha_efecto4 else ''),
            'F_Hoy': f_hoy,
            'F_OfPres': folio,
            'COMENTARIOS': observaciones,

            'Nombre_Interino': nombre_interino,
            'CURP_Interino': curp_interino,
            'RFC_Interino': rfc_interino,
            'Dom_Particular': domicilio_part_interino,
            'C_P_Interino': codigo_postal_interino,
            'Poblacion_Interino': poblacion_interino,
            'Telefono_Interino': telefono_interino,
            'Presupuestal_Interino': presupuestal_interino,
            'Funcion_Interino': funcion_interino,
            'Tipo_Movimiento_Interino': tipo_movimiento_interino,

            'Codigo_Interino': codigo_interino,

            'Paterno': paterno_interino,
            'Materno': materno_interino,
            'Nombre': nombre_interino_solo,
            'Formacion_Academica': formacion_academica_interino,

            # Prelacion Data (ahora se llenan automáticamente)
            'No_Prel': no_prel,
            'Folio_Prel': folio_prel,
            'Tipo_Val': tipo_val,

            # Autoridades
            'Supervisor': supervisor_adscripcion_info['nombre'],
            'P_Sup': supervisor_adscripcion_info['nivel'],
            'Director': director_adscripcion_info['nombre'],
            'P_Dir': director_adscripcion_info['nivel'],

            # --- NUEVAS Variables para Autoridades de la Escuela de Pago ---
            'Supervisor_Techo_F': supervisor_pago_info['nombre'],
            'P_Sup_Techo_F': supervisor_pago_info['nivel'],
            'Director_Techo_F': director_pago_info['nombre'],
            'P_Dir_Techo_F': director_pago_info['nivel'],

            # VBA specific fields
            'Resultado_Alta': tipo_movimiento_interino,
            'QuincenaInicial': '',
            'QuincenaFinal': '',
            'Horario': maestro_titular.horario if maestro_titular else '',
            'TipoPlaza': 'JORNADA' if (maestro_titular and maestro_titular.hrs == "00.0") else "HORA/SEMANA/MES",
            'Horas': maestro_titular.hrs.split('.')[0] if (maestro_titular and maestro_titular.hrs and '.' in maestro_titular.hrs) else '',
            'Nivel': 'Educación Especial',
            'Entidad': 'DURANGO',
            'Municipio': escuela_adscripcion_info['region'],
            'Region': escuela_adscripcion_info['region'],
            'ZonaEconomica': escuela_adscripcion_info['zona_economica'],
            'Destino': '',
            'Apreciacion': '',
            'TipoVacante': '',
            'NoOrdenamiento': '',
            'FolioOrdenamiento': '',
            'CurpInterino': curp_interino,
            'NombreInterino': nombre_interino,
            'Tipo': motivo_movimiento,
            'Observaciones': observaciones,
            'QuincenaInicio': quincena_inicial,
            'QuincenaFinal': quincena_final,

            'I_Dia': i_dia,
            'I_Mes': i_mes,
            'I_Ano': i_ano,
            'F_Dia': f_dia,
            'F_Mes': f_mes,
            'F_Ano': f_ano,

            'F_HoyLetra': f_hoy_letras,
        }

        # --- Render Document ---
        doc.render(context)

        # --- Save Document ---
        output_base_dir = os.path.join(settings.BASE_DIR, 'tramites_generados')
        template_name_clean = plantilla_tramite.nombre.replace(" ", "_").replace(".", "").replace("(", "").replace(")", "").replace(",", "").replace("-", "").upper()

        subfolder_map = {
            "REINGRESO": "reingresos",
            "FILIACION": "filiacion",
            "SOLICITUD_DE_ASIGNACION": "solicitud_asignacion",
            "REINGRESO_SIN_PRELACION": "reingreso_sin_prelacion",
            "JUSTIFICACION_DE_PERFIL": "justificacion_perfil",
            "REPORTE_DE_VACANCIA": "reporte_vacancia",
            "CONSTANCIAS": "constancias",
            "CAMBIO_DEL_CENTRO_DE_TRABAJO": "cambio_ct",
            "CUADRO_CAMBIOS_CON_FOLIO": "cuadro_cambios",
            "PROPUESTA_DE_MOVIMIENTO": "propuesta_movimiento",
            "OFICIO_DE_REINCORPORACION": "oficio_reincorporacion",
        }
        subfolder = subfolder_map.get(template_name_clean, "otros_tramites")

        output_dir = os.path.join(output_base_dir, subfolder)
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"TRAMITE_{template_name_clean}_{timestamp}.docx"
        output_path = os.path.join(output_dir, output_filename)

        doc.save(output_path)
        return True, output_path
    except Exception as e:
        print(f"Error generating Word document: {e}")
        return False, str(e)

# Vista AJAX para obtener datos de prelación
def get_prelacion_data_ajax(request):
    curp_interino = request.GET.get('curp_interino')
    datos_prelacion = {
        'encontrado': False,
        'numero_prelacion': '',
        'folio_prelacion': '',
        'tipo_val': '',
        'nombre_prelacion': ''
    }

    if curp_interino:
        try:
            # Buscar en la tabla Prelacion por la CURP del interino
            prelacion = Prelacion.objects.filter(curp=curp_interino).first()
            
            if prelacion:
                datos_prelacion = {
                    'encontrado': True,
                    'numero_prelacion': prelacion.pos_orden or '',
                    'folio_prelacion': prelacion.folio or '',
                    'tipo_val': prelacion.tipo_val or '',
                    'nombre_prelacion': prelacion.nombre or ''
                }
        except Exception as e:
            print(f"Error buscando prelación: {e}")

    return JsonResponse(datos_prelacion)

# Vistas principales
@login_required
def index(request):
    context = {
        'titulo': 'Dashboard'
    }

    # Check permissions for each section
    if request.user.has_perm('gestion_escolar.ver_estadisticas_generales'):
        context['total_zonas'] = Zona.objects.count()
        context['total_escuelas'] = Escuela.objects.count()
        context['total_maestros'] = Maestro.objects.count()
        context['total_directores'] = Maestro.objects.filter(funcion__icontains='DIRECTOR').count()

    if request.user.has_perm('gestion_escolar.ver_grafico_distribucion_zona'):
        distribucion_por_zona = Zona.objects.annotate(num_escuelas=Count('escuela')).order_by('numero')
        context['zona_labels'] = json.dumps([f"Zona {zona.numero}" for zona in distribucion_por_zona])
        context['zona_data'] = json.dumps([zona.num_escuelas for zona in distribucion_por_zona])

    if request.user.has_perm('gestion_escolar.ver_lista_pendientes'):
        today = timezone.now().date()
        context['ultimos_pendientes'] = Pendiente.objects.filter(
            usuario=request.user,
            completado=False,
            fecha_programada__lte=today
        ).order_by('fecha_programada')[:5]

    if request.user.has_perm('gestion_escolar.ver_lista_ultimo_personal'):
        context['ultimo_personal'] = Maestro.objects.order_by('-fecha_registro')[:5]

    if request.user.has_perm('gestion_escolar.view_registrocorrespondencia'):
        context['ultima_correspondencia'] = RegistroCorrespondencia.objects.order_by('-fecha_registro')[:5]

    return render(request, 'gestion_escolar/index.html', context)

@permission_required('gestion_escolar.acceder_ajustes', raise_exception=True)
def ajustes_view(request):
    return render(request, 'gestion_escolar/ajustes.html', {'titulo': 'Ajustes'})


@login_required
def cambiar_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Tu contraseña ha sido actualizada correctamente.')
            return redirect('ajustes')
        else:
            messages.error(request, 'Por favor corrige los errores.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'gestion_escolar/cambiar_password.html', {
        'form': form,
        'titulo': 'Cambiar Contraseña'
    })

@login_required
def editar_perfil(request):
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tu perfil ha sido actualizado correctamente.')
            return redirect('ajustes')
    else:
        form = CustomUserChangeForm(instance=request.user)
    return render(request, 'gestion_escolar/editar_perfil.html', {'form': form, 'titulo': 'Editar Perfil'})


@user_passes_test(lambda u: u.is_superuser)
def asignar_director(request):
    if request.method == 'POST':
        form = AsignarDirectorForm(request.POST)
        if form.is_valid():
            maestro = form.cleaned_data['maestro']
            user = form.cleaned_data['usuario']
            
            maestro.user = user
            maestro.save()
            
            directores_group, created = Group.objects.get_or_create(name='Directores')
            user.groups.add(directores_group)
            
            messages.success(request, f'El maestro {maestro} ha sido asignado como director al usuario {user}.')
            return redirect('ajustes')
    else:
        form = AsignarDirectorForm()
    return render(request, 'gestion_escolar/asignar_director.html', {'form': form, 'titulo': 'Asignar Director'})


# Vistas para Zonas
def lista_zonas(request):
    if request.user.groups.filter(name='Directores').exists():
        raise PermissionDenied
    
    # Usamos select_related para traer el supervisor en la misma consulta
    zonas = Zona.objects.select_related('supervisor').order_by('numero')

    return render(request, 'gestion_escolar/lista_zonas.html', {'zonas': zonas})

def agregar_zona(request):
    if request.method == 'POST':
        form = ZonaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Zona agregada correctamente.')
            return redirect('lista_zonas')
        else:
            messages.error(request, 'Por favor corrige los errores.')
    else:
        form = ZonaForm()
    return render(request, 'gestion_escolar/form_zona.html', {'form': form, 'titulo': 'Agregar Zona'})

def editar_zona(request, pk):
    zona = get_object_or_404(Zona.objects.select_related('supervisor'), pk=pk)
    if request.method == 'POST':
        form = ZonaForm(request.POST, instance=zona)
        if form.is_valid():
            form.save()
            messages.success(request, 'Zona actualizada correctamente.')
            return redirect('lista_zonas')
        else:
            messages.error(request, 'Por favor corrige los errores.')
    else:
        form = ZonaForm(instance=zona)
    
    return render(request, 'gestion_escolar/form_zona.html', {
        'form': form, 
        'zona': zona, 
        'titulo': 'Editar Zona'
    })

def eliminar_zona(request, pk):
    zona = get_object_or_404(Zona, pk=pk)
    if request.method == 'POST':
        zona.delete()
        messages.success(request, 'Zona eliminada correctamente.')
        return redirect('lista_zonas')
    return render(request, 'gestion_escolar/eliminar_zona.html', {'zona': zona})

def detalle_zona(request, pk):
    # Usamos select_related para optimizar y traer el supervisor en la misma consulta
    zona = get_object_or_404(Zona.objects.select_related('supervisor'), pk=pk)
    
    # Obtenemos todas las escuelas que pertenecen a esta zona
    escuelas_en_zona = Escuela.objects.filter(zona_esc=zona).order_by('nombre_ct')
    
    context = {
        'zona': zona,
        'escuelas': escuelas_en_zona,
        'titulo': f"Detalle de la Zona {zona.numero}"
    }
    
    return render(request, 'gestion_escolar/detalle_zona.html', context)


# Vistas para Escuelas
def lista_escuelas(request):
    if request.user.groups.filter(name='Directores').exists():
        raise PermissionDenied
    escuelas = Escuela.objects.all().order_by('nombre_ct')
    return render(request, 'gestion_escolar/lista_escuelas.html', {'escuelas': escuelas})

def agregar_escuela(request):
    if request.method == 'POST':
        form = EscuelaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Escuela agregada correctamente.')
            return redirect('lista_escuelas')
        else:
            messages.error(request, 'Por favor corrige los errores.')
    else:
        form = EscuelaForm()
    return render(request, 'gestion_escolar/form_escuela.html', {'form': form, 'titulo': 'Agregar Escuela'})

def editar_escuela(request, pk):
    escuela = get_object_or_404(Escuela, pk=pk)
    if request.method == 'POST':
        form = EscuelaForm(request.POST, instance=escuela)
        if form.is_valid():
            form.save()
            messages.success(request, 'Escuela actualizada correctamente.')
            return redirect('lista_escuelas')
        else:
            messages.error(request, 'Por favor corrige los errores.')
    else:
        form = EscuelaForm(instance=escuela)
    return render(request, 'gestion_escolar/form_escuela.html', {'form': form, 'titulo': 'Editar Escuela'})

def eliminar_escuela(request, pk):
    escuela = get_object_or_404(Escuela, pk=pk)
    if request.method == 'POST':
        escuela.delete()
        messages.success(request, 'Escuela eliminada correctamente.')
        return redirect('lista_escuelas')
    return render(request, 'gestion_escolar/eliminar_escuela.html', {'escuela': escuela})

def detalle_escuela(request, pk):
    escuela = get_object_or_404(Escuela, pk=pk)
    personal = Maestro.objects.filter(id_escuela=escuela)
    context = {
        'escuela': escuela,
        'personal': personal,
        'titulo': 'Detalle de la Escuela'
    }
    return render(request, 'gestion_escolar/detalle_escuela.html', context)

# Vistas para Maestros
from unidecode import unidecode
from django.urls import reverse

@login_required
def lista_maestros(request):
    user = request.user
    # Esta vista ahora solo renderiza el template base.
    return render(request, 'gestion_escolar/lista_maestros.html')

@login_required
def lista_maestros_ajax(request):
    # Parámetros de DataTables
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    # Ordenamiento
    order_column_index = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'asc')
    column_names = ['id_maestro', 'a_paterno', 'id_escuela__id_escuela', 'curp', 'clave_presupuestal', 'status']
    order_column = column_names[order_column_index]
    if order_dir == 'desc':
        order_column = f'-{order_column}'

    # Queryset base
    user = request.user
    if user.groups.filter(name='Directores').exists():
        try:
            maestro_director = user.maestro_profile
            queryset = Maestro.objects.filter(id_escuela=maestro_director.id_escuela)
        except AttributeError:
            queryset = Maestro.objects.none()
    else:
        queryset = Maestro.objects.all()
    
    # Optimización: Precargar los datos de la escuela para evitar consultas N+1
    queryset = queryset.select_related('id_escuela')

    queryset = queryset.exclude(id_maestro__isnull=True).exclude(id_maestro='')

    # Total de registros sin filtrar
    records_total = queryset.count()

    # Filtro de búsqueda
    if search_value:
        search_terms = search_value.split()
        queries = []
        for term in search_terms:
            # Construimos una consulta Q para cada término de búsqueda
            # Esto busca el término en cualquiera de los campos especificados
            term_query = Q(
                Q(id_maestro__icontains=term) |
                Q(nombres__icontains=term) |
                Q(a_paterno__icontains=term) |
                Q(a_materno__icontains=term) |
                Q(curp__icontains=term) |
                Q(clave_presupuestal__icontains=term) |
                Q(id_escuela__id_escuela__icontains=term) # Búsqueda por C.C.T.
            )
            queries.append(term_query)
        # Combinamos todas las consultas con un operador AND
        queryset = queryset.filter(*queries)

    # Total de registros después del filtro
    records_filtered = queryset.count()

    # Ordenamiento y paginación
    queryset = queryset.order_by(order_column)[start:start + length]

    # Preparar datos para la respuesta JSON
    data = []
    for maestro in queryset:
        actions = '<div class="btn-group" role="group">'
        # View button is always available if they can see the list
        actions += f'<a href="{reverse('detalle_maestro', args=[maestro.pk])}" class="btn btn-sm btn-outline-info"><i class="fas fa-eye"></i></a>'
        
        # Conditional Edit button
        if request.user.has_perm('gestion_escolar.change_maestro'):
            actions += f'<a href="{reverse('editar_maestro', args=[maestro.pk])}" class="btn btn-sm btn-outline-primary"><i class="fas fa-edit"></i></a>'

        # Conditional Delete button
        if request.user.has_perm('gestion_escolar.delete_maestro'):
            actions += f'<a href="{reverse('eliminar_maestro', args=[maestro.pk])}" class="btn btn-sm btn-outline-danger"><i class="fas fa-trash"></i></a>'
        
        actions += '</div>'
        status_map = {'ACTIVO': 'success', 'INACTIVO': 'warning'}
        status_class = status_map.get(maestro.status, 'secondary')
        status_html = f'<span class="badge bg-{status_class}">{maestro.get_status_display()}</span>'

        # Comprobar si el CCT de la escuela es diferente del techo financiero
        is_misplaced = False
        if maestro.id_escuela and maestro.techo_f:
            if maestro.id_escuela.id_escuela.strip().upper() != maestro.techo_f.strip().upper():
                is_misplaced = True

        data.append([
            maestro.id_maestro,
            f'{maestro.a_paterno} {maestro.a_materno} {maestro.nombres}',
            maestro.id_escuela.id_escuela if maestro.id_escuela else 'N/A',
            maestro.curp,
            maestro.clave_presupuestal or '-',
            status_html,
            actions,
            is_misplaced  # Dato extra para el frontend
        ])

    response = {
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data,
    }
    return JsonResponse(response)
def agregar_maestro(request):
    all_escuelas = Escuela.objects.all()
    initial_data = {}
    escuela_id = request.GET.get('escuela_id')
    if escuela_id:
        try:
            escuela = Escuela.objects.get(pk=escuela_id)
            initial_data['id_escuela'] = escuela
        except Escuela.DoesNotExist:
            pass # La escuela no existe, el formulario se inicializará sin ella

    if request.method == 'POST':
        form = MaestroForm(request.POST, request=request)
        if form.is_valid():
            maestro = form.save(commit=False)
            maestro.save()
            messages.success(request, 'Maestro agregado correctamente.')
            # Redirigir de vuelta al detalle de la escuela si se agregó desde allí
            if escuela_id:
                return redirect('detalle_escuela', pk=escuela_id)
            return redirect('lista_maestros')
        else:
            messages.error(request, 'Por favor corrige los errores.')
    else:
        form = MaestroForm(initial=initial_data, request=request)
    return render(request, 'gestion_escolar/form_maestro.html', {
        'form': form,
        'titulo': 'Agregar Maestro',
        'all_escuelas': all_escuelas
    })

@login_required
def editar_maestro(request, pk):
    maestro = get_object_or_404(Maestro, id_maestro=pk)
    documentos = maestro.documentos_expediente.all()
    
    if request.method == 'POST':
        # Distinguir si se está subiendo un documento o guardando el maestro
        if 'submit_documento' in request.POST:
            doc_form = DocumentoExpedienteForm(request.POST, request.FILES)
            if doc_form.is_valid():
                documento = doc_form.save(commit=False)
                documento.maestro = maestro
                documento.subido_por = request.user
                documento.save()
                messages.success(request, 'Documento subido correctamente.')
                return redirect('editar_maestro', pk=maestro.id_maestro)
            else:
                messages.error(request, 'Error al subir el documento.')
            # Si el formulario de documento falla, necesitamos el formulario principal para renderizar la página
            form = MaestroForm(instance=maestro, request=request)
        else:
            form = MaestroForm(request.POST, instance=maestro, request=request)
            if form.is_valid():
                form.save()
                messages.success(request, 'Maestro actualizado correctamente.')
                return redirect('lista_maestros')
            else:
                messages.error(request, 'Por favor corrige los errores.')
            # Si el formulario principal falla, necesitamos el formulario de documento vacío
            doc_form = DocumentoExpedienteForm()
    else:
        form = MaestroForm(instance=maestro, request=request)
        doc_form = DocumentoExpedienteForm()

    context = {
        'form': form,
        'doc_form': doc_form,
        'maestro': maestro,
        'documentos': documentos,
        'titulo': 'Editar Maestro'
    }
    return render(request, 'gestion_escolar/form_maestro.html', context)

def eliminar_maestro(request, pk):
    maestro = get_object_or_404(Maestro, id_maestro=pk)
    if request.method == 'POST':
        maestro.delete()
        messages.success(request, 'Maestro eliminada correctamente.')
        return redirect('lista_maestros')
    return render(request, 'gestion_escolar/eliminar_maestro.html', {'maestro': maestro})

def detalle_maestro(request, pk):
    maestro = get_object_or_404(Maestro, id_maestro=pk)
    documentos = maestro.documentos_expediente.all()
    
    if request.method == 'POST':
        form = DocumentoExpedienteForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.maestro = maestro
            documento.subido_por = request.user
            documento.save()
            messages.success(request, 'Documento subido correctamente.')
            return redirect('detalle_maestro', pk=maestro.id_maestro)
        else:
            messages.error(request, 'Error al subir el documento. Por favor, revise el formulario.')
    else:
        form = DocumentoExpedienteForm()

    context = {
        'maestro': maestro,
        'documentos': documentos,
        'form': form,
        'titulo': 'Detalle del Personal'
    }
    return render(request, 'gestion_escolar/detalle_maestro.html', context)

@login_required
def eliminar_documento_expediente(request, doc_pk):
    documento = get_object_or_404(DocumentoExpediente, pk=doc_pk)
    maestro_id = documento.maestro.id_maestro

    if request.method == 'POST':
        try:
            # Delete the file from storage
            documento.archivo.delete(save=False)
            # Delete the model instance
            documento.delete()
            messages.success(request, 'Documento eliminado correctamente.')
            return redirect('detalle_maestro', pk=maestro_id)
        except Exception as e:
            messages.error(request, f"Error al eliminar el documento: {e}")
            return redirect('detalle_maestro', pk=maestro_id)

    context = {
        'documento': documento,
        'maestro': documento.maestro,
    }
    return render(request, 'gestion_escolar/eliminar_documento_expediente.html', context)

# Vistas para diferentes funciones
def lista_por_funcion(request, funcion):
    # Mapeo de términos de búsqueda a nombres para mostrar y sus posibles valores en la BD
    funcion_mapping = {
        'DIRECTOR': {'display': 'Director', 'values': ['DIRECTOR', 'DIRECTOR (A)']},
        'SUPERVISOR': {'display': 'Supervisor', 'values': ['SUPERVISOR', 'SUPERVISOR (A)', 'SUPERVISOR(A)']},
        'MAESTRO_GRUPO': {'display': 'Maestro de Grupo', 'values': ['MAESTRO_GRUPO', 'MAESTRO(A) DE GRUPO', 'MAESTRO(A) DE GRUPO CON ESPECIALIDAD', 'MAESTRO(A) DE GRUPO ESPECIALISTA','MATRO(A) DE GRUPO ESPECIALISTA']},
        'DOCENTE_APOYO': {'display': 'Docente de Apoyo', 'values': ['MAESTRO(A) DE APOYO']},
        'PSICOLOGO': {'display': 'Psicólogo', 'values': ['PSICOLOGO', 'PSICÓLOGO(A)', 'PSICÓLOGO (A)']},
        'TRABAJADOR_SOCIAL': {'display': 'Trabajador Social', 'values': ['TRABAJADOR_SOCIAL', 'TRABAJADOR (A) SOCIAL']},
        'NIÑERO': {'display': 'Niñero', 'values': ['NIÑERO', 'NIÑERO(A)']},
        'SECRETARIO': {'display': 'Secretario', 'values': ['SECRETARIO', 'SECRETARIA ', 'SECRETARIO(A)']},
        'INTENDENTE': {'display': 'Intendente', 'values': ['INTENDENTE']},
        'VELADOR': {'display': 'Velador', 'values': ['VELADOR']},
        'VIGILANTE': {'display': 'Vigilante', 'values': ['VIGILANTE', 'VIGILANTE ']},
        'OTRO': {'display': 'Otro', 'values': ['OTRO']},
        'APOYO_TECNICO_PEDAGOGICO': {'display': 'Apoyo Técnico Pedagógico', 'values': ['APOYO TECNICO PEDAGOGICO']},
        'INSTRUCTOR_TALLER': {'display': 'Instructor de Taller', 'values': ['INSTRUCTOR(A) DE TALLER']},
        'MAESTRO_TALLER': {'display': 'Maestro de Taller', 'values': ['MAESTRO(A) DE TALLER', 'MAESTRO DE TALLER']},
        'MAESTRO_MUSICA': {'display': 'Maestro de Música', 'values': ['MAESTRO(A) MUSICA']},
        'MAESTRO_EDUCACION_FISICA': {'display': 'Maestro de Educación Física', 'values': ['MAESTRO(A) DE EDUCACIÓN FÍSICA']},
        'MEDICO': {'display': 'Médico', 'values': ['MÉDICO(A)', 'MÉDICO (A)']},
        'PROMOTOR_TIC': {'display': 'Promotor TIC', 'values': ['PROMOTOR TIC', 'PROMOTOR TIC ']},
        'TERAPISTA_FISICO': {'display': 'Terapista Físico', 'values': ['TERAPISTA FISICO ']},
        'BIBLIOTECARIO': {'display': 'Bibliotecario', 'values': ['BIBLIOTECARIO(A)']},
        'ADMINISTRATIVO_ESPECIALIZADO': {'display': 'Administrativo Especializado', 'values': ['ADMINISTRATIVO ESPECIALIZADO']},
        'OFICIAL_SERVICIOS_MANTENIMIENTO': {'display': 'Oficial de Servicios y Mantenimiento', 'values': ['OFICIAL DE SERVICIOS Y MANTENIMIENTO', 'OFICIAL DE SERVICIOS DE MANTENIMIENTO']},
        'ASISTENTE_DE_SERVICIOS': {'display': 'Asistente de Servicios', 'values': ['ASISTENTE DE SERVICIOS']},
        'ASESOR_JURIDICO': {'display': 'Asesor Jurídico', 'values': ['ASESOR JURÍDICO']},
        'AUXILIAR_DE_GRUPO': {'display': 'Auxiliar de Grupo', 'values': ['AUXILIAR DE GRUPO']},
        'MAESTRO_COMUNICACION': {'display': 'Maestro de Comunicación', 'values': ['MAESTRO(A) DE COMUNICACIÓN ']},
        'MAESTRO_AULA_HOSPITALARIA': {'display': 'Maestro Aula Hospitalaria', 'values': ['MAESTRO(A) AULA HOSPITALARIA']},
    }

    # Obtener los valores de función a buscar
    funcion_info = funcion_mapping.get(funcion)
    if not funcion_info:
        # Si la función no está en el mapeo, intentar con el nombre tal cual
        funcion_values = [funcion]
        funcion_display = funcion.replace(' ', '_').title()
    else:
        funcion_values = funcion_info['values']
        funcion_display = funcion_info['display']

    print(f"DEBUG: Buscando maestros con funcion__in: {funcion_values}")

    trabajadores = Maestro.objects.filter(
        funcion__in=funcion_values
    ).exclude(
        id_maestro__isnull=True
    ).exclude(
        id_maestro=''
    ).order_by('a_paterno', 'a_materno', 'nombres')

    print(f"DEBUG: Se encontraron {trabajadores.count()} maestros.")

    return render(request, 'gestion_escolar/lista_por_funcion.html', {
        'trabajadores': trabajadores,
        'funcion': funcion,
        'funcion_display': funcion_display,
        'titulo': f'{funcion_display}es' if funcion.endswith('OR') else f'{funcion_display}s'
    })

# Vistas específicas para cada función (opcional pero útil para URLs amigables)
def lista_directores(request):
    return lista_por_funcion(request, 'DIRECTOR')

def lista_supervisores_maestros(request):
    return lista_por_funcion(request, 'SUPERVISOR')

def lista_maestros_grupo(request):
    return lista_por_funcion(request, 'MAESTRO_GRUPO')

def lista_psicologos(request):
    return lista_por_funcion(request, 'PSICOLOGO')

def lista_trabajadores_sociales(request):
    return lista_por_funcion(request, 'TRABAJADOR_SOCIAL')

def lista_docentes_apoyo(request):
    return lista_por_funcion(request, 'DOCENTE_APOYO')

# Vistas para Trámites
@permission_required('gestion_escolar.acceder_tramites', raise_exception=True)
def generar_tramites_generales(request):
    if request.method == 'POST':
        form = TramiteForm(request.POST, form_type='tramites') # Pass form_type
        if form.is_valid():
            plantilla_id = form.cleaned_data['plantilla'].id
            plantilla_tramite = PlantillaTramite.objects.get(id=plantilla_id)

            success, message = generate_word_document(form.cleaned_data, plantilla_tramite, request.user)

            if success:
                # Crear registro en el historial
                try:
                    # --- ENRIQUECER DATOS PARA EL HISTORIAL ---
                    # Copiamos los datos limpios del formulario
                    datos_para_historial = form.cleaned_data.copy()
                    
                    # Obtenemos los objetos relacionados
                    maestro_titular_obj = form.cleaned_data.get('maestro_titular')
                    escuela_titular = maestro_titular_obj.id_escuela if maestro_titular_obj else None
                    zona_esc = escuela_titular.zona_esc if escuela_titular else None

                    # Usamos las funciones de ayuda para obtener la información
                    escuela_info = get_school_info(escuela_titular)
                    director_info = get_director_info(escuela_titular)
                    supervisor_info = get_supervisor_info(zona_esc)

                    # Agregamos la información adicional al diccionario que se guardará
                    datos_para_historial['techo_financiero_titular'] = maestro_titular_obj.techo_f if maestro_titular_obj else ''
                    datos_para_historial['clave_ct'] = escuela_info.get('id_escuela', '')
                    datos_para_historial['nombre_ct'] = escuela_info.get('nombre_ct', '')
                    datos_para_historial['turno'] = escuela_info.get('turno', '')
                    datos_para_historial['domicilio_ct'] = escuela_info.get('domicilio', '')
                    datos_para_historial['z_escolar'] = escuela_info.get('zona_esc_numero', '')
                    datos_para_historial['region'] = escuela_info.get('region', '')
                    datos_para_historial['sostenimiento'] = escuela_info.get('sostenimiento', '')
                    datos_para_historial['supervisor'] = supervisor_info.get('nombre', '')
                    datos_para_historial['director'] = director_info.get('nombre', '')
                    
                    print(f"DEBUG: Maestro Titular en generar_tramites_generales: {maestro_titular_obj}") # DEBUG LINE
                    Historial.objects.create(
                        usuario=request.user,
                        tipo_documento=f"Trámite - {plantilla_tramite.nombre}",
                        maestro=maestro_titular_obj,
                        ruta_archivo=message,
                        motivo=form.cleaned_data.get('motivo_tramite').motivo_tramite if form.cleaned_data.get('motivo_tramite') else '',
                        maestro_secundario_nombre=get_full_name(form.cleaned_data.get('maestro_interino')),
                        datos_tramite=serialize_form_data(datos_para_historial) # Guardamos los datos enriquecidos
                    )
                except Exception as e:
                    messages.warning(request, f"Advertencia: El trámite se generó pero no se pudo guardar en el historial: {e}")

                # Leer el archivo generado y enviarlo como respuesta para descarga
                with open(message, 'rb') as doc_file:
                    response = HttpResponse(doc_file.read(), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
                    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(message)}"'
                    messages.success(request, 'Trámite generado y descargado correctamente.')
                    return response
            else:
                messages.error(request, f'Error al generar el trámite: {message}')
                return redirect('generar_tramites_generales') # Redirigir en caso de error
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = TramiteForm(form_type='tramites') # Pass form_type

    context = {
        'form': form,
        'titulo': 'Generar Trámite'
    }
    return render(request, 'gestion_escolar/generar_tramite.html', context)

@permission_required('gestion_escolar.acceder_oficios', raise_exception=True)
def generar_oficios(request):
    if request.method == 'POST':
        form = TramiteForm(request.POST, form_type='oficios') # Pass form_type
        if form.is_valid():
            plantilla_id = form.cleaned_data['plantilla'].id
            plantilla_tramite = PlantillaTramite.objects.get(id=plantilla_id)

            success, message = generate_word_document(form.cleaned_data, plantilla_tramite, request.user)

            if success:
                # Crear registro en el historial
                try:
                    # --- ENRIQUECER DATOS PARA EL HISTORIAL (para Oficios) ---
                    datos_para_historial = form.cleaned_data.copy()
                    
                    maestro_titular_obj = form.cleaned_data.get('maestro_titular')
                    escuela_titular = maestro_titular_obj.id_escuela if maestro_titular_obj else None
                    zona_esc = escuela_titular.zona_esc if escuela_titular else None

                    escuela_info = get_school_info(escuela_titular)
                    director_info = get_director_info(escuela_titular)
                    supervisor_info = get_supervisor_info(zona_esc)

                    datos_para_historial['techo_financiero_titular'] = maestro_titular_obj.techo_f if maestro_titular_obj else ''
                    datos_para_historial['clave_ct'] = escuela_info.get('id_escuela', '')
                    datos_para_historial['nombre_ct'] = escuela_info.get('nombre_ct', '')
                    datos_para_historial['turno'] = escuela_info.get('turno', '')
                    datos_para_historial['domicilio_ct'] = escuela_info.get('domicilio', '')
                    datos_para_historial['z_escolar'] = escuela_info.get('zona_esc_numero', '')
                    datos_para_historial['region'] = escuela_info.get('region', '')
                    datos_para_historial['sostenimiento'] = escuela_info.get('sostenimiento', '')
                    datos_para_historial['supervisor'] = supervisor_info.get('nombre', '')
                    datos_para_historial['director'] = director_info.get('nombre', '')

                    Historial.objects.create(
                        usuario=request.user,
                        tipo_documento=f"Oficio - {plantilla_tramite.nombre}",
                        maestro=form.cleaned_data.get('maestro_titular'),
                        ruta_archivo=message,
                        motivo=form.cleaned_data.get('motivo_tramite').motivo_tramite if form.cleaned_data.get('motivo_tramite') else '',
                        maestro_secundario_nombre=get_full_name(form.cleaned_data.get('maestro_interino')),
                        datos_tramite=serialize_form_data(datos_para_historial)
                    )
                except Exception as e:
                    messages.warning(request, f"Advertencia: El oficio se generó pero no se pudo guardar en el historial: {e}")

                # Leer el archivo generado y enviarlo como respuesta para descarga
                with open(message, 'rb') as doc_file:
                    response = HttpResponse(doc_file.read(), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
                    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(message)}"'
                    messages.success(request, 'Oficio generado y descargado correctamente.')
                    return response
            else:
                messages.error(request, f'Error al generar el oficio: {message}')
                return redirect('generar_oficios') # Redirigir en caso de error
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = TramiteForm(form_type='oficios') # Pass form_type

    context = {
        'form': form,
        'titulo': 'Generar Oficio'
    }
    return render(request, 'gestion_escolar/generar_tramite.html', context)

def get_motivos_tramite_ajax(request):
    plantilla_id = request.GET.get('plantilla_id')
    motivos_filtrados = []

    if plantilla_id:
        try:
            plantilla = PlantillaTramite.objects.get(id=plantilla_id)
            opcion = plantilla.nombre.strip().upper() # Normalizar el nombre de la plantilla

            if opcion == "REINGRESO" or opcion == "FILIACION" or opcion == "SOLICITUD DE ASIGNACION" or opcion == "REINGRESO SIN PRELACION" or opcion == "JUSTIFICACION DE PERFIL" or opcion == "REPORTE DE VACANCIA":
                ids = [1, 2, 3, 4, 5, 6, 7, 21, 22, 24, 38]
            elif opcion == "CONSTANCIAS":
                ids = [20, 23, 29, 30, 31, 32, 33, 34, 35, 36, 37]
            elif opcion == "CAMBIO DEL CENTRO DE TRABAJO":
                ids = [19, 13]
            elif opcion == "CUADRO CAMBIOS CON FOLIO":
                ids = [13]
            elif opcion == "PROPUESTA DE MOVIMIENTO":
                ids = [11, 12, 25, 26, 27]
            elif opcion == "ALTA INICIAL": # Nuevo tipo de trámite
                ids = [39]
            elif opcion == "OFICIO DE REINCORPORACION":
                ids = [1, 2, 4, 15, 21, 22, 24]
            else:
                ids = []

            if ids:
                motivos_filtrados_qs = MotivoTramite.objects.filter(id__in=ids).order_by('motivo_tramite')
            else:
                motivos_filtrados_qs = MotivoTramite.objects.all().order_by('motivo_tramite')

            for motivo in motivos_filtrados_qs:
                motivos_filtrados.append({'id': motivo.id, 'text': motivo.motivo_tramite})

        except PlantillaTramite.DoesNotExist:
            pass

    return JsonResponse(motivos_filtrados, safe=False)

@login_required
def buscar_maestros_ajax(request):
    search_term = request.GET.get('term', '')
    
    if not search_term or len(search_term) < 2:
        return JsonResponse({'results': []})
    
    terms = [term for term in search_term.split() if term]
    query = Q()
    
    for term in terms:
        query &= (
            Q(nombres__icontains=term) |
            Q(a_paterno__icontains=term) |
            Q(a_materno__icontains=term)
        )
    
    maestros = Maestro.objects.filter(query).order_by('a_paterno', 'a_materno', 'nombres')[:20]
    
    results = []
    for maestro in maestros:
        full_name = f"{maestro.a_paterno or ''} {maestro.a_materno or ''} {maestro.nombres or ''}".strip()
        results.append({
            "id": maestro.id_maestro,
            "text": full_name
        })
    
    return JsonResponse({'results': results})

@login_required
def get_maestro_data_ajax(request):
    maestro_id = request.GET.get('maestro_id')

    data = {}
    if maestro_id:
        try:
            maestro = Maestro.objects.get(id_maestro=maestro_id)
            data = {
                'curp': maestro.curp or '',
                'rfc': maestro.rfc or '',
                'clave_presupuestal': maestro.clave_presupuestal or '',
                'categoria': maestro.categog.descripcion if maestro.categog else '',
                'funcion': maestro.funcion or '',
            }
        except Maestro.DoesNotExist:
            data = {'error': 'Maestro no encontrado'}
    return JsonResponse(data)

# Vistas para Categorías
def lista_categorias(request):
    if request.user.groups.filter(name='Directores').exists():
        raise PermissionDenied
    query = request.GET.get('q')
    categorias = Categoria.objects.all().order_by('id_categoria')

    if query:
        categorias = categorias.filter(
            Q(id_categoria__icontains=query) | Q(descripcion__icontains=query)
        )

    context = {
        'categorias': categorias,
        'query': query,
    }
    return render(request, 'gestion_escolar/lista_categorias.html', context)

def editar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada correctamente.')
            return redirect('lista_categorias')
        else:
            messages.error(request, 'Por favor corrige los errores.')
    else:
        form = CategoriaForm(instance=categoria)
    return render(request, 'gestion_escolar/form_categoria.html', {'form': form, 'titulo': 'Editar Categoría'})

def eliminar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'Categoría eliminada correctamente.')
        return redirect('lista_categorias')
    return render(request, 'gestion_escolar/eliminar_categoria.html', {'categoria': categoria})

from django.http import JsonResponse, HttpResponse, FileResponse, HttpResponseForbidden
from django.db.models.functions import Upper, Trim

@permission_required('gestion_escolar.acceder_reportes', raise_exception=True)
def reportes_dashboard(request):
    context = {
        'titulo': 'Dashboard de Reportes'
    }
    return render(request, 'gestion_escolar/reportes_dashboard.html', context)

@login_required
def reporte_personal_fuera_adscripcion(request):
    # Anotamos versiones limpias (sin espacios y en mayúsculas) de los campos a comparar
    personal_qs = Maestro.objects.annotate(
        techo_f_clean=Trim(Upper('techo_f')),
        id_escuela_clean=Trim(Upper('id_escuela__id_escuela'))
    ).exclude(techo_f__isnull=True).exclude(techo_f='')

    # Filtramos donde los campos limpios no coinciden
    personal_fuera_adscripcion = [p for p in personal_qs if p.techo_f_clean != p.id_escuela_clean]

    context = {
        'personal_fuera_adscripcion': personal_fuera_adscripcion,
        'titulo': 'Reporte de Personal Fuera de Adscripción'
    }
    return render(request, 'gestion_escolar/reporte_fuera_adscripcion.html', context)

@permission_required('gestion_escolar.acceder_historial', raise_exception=True)
def historial(request):
    historial_items = Historial.objects.select_related('usuario', 'maestro').all().order_by('-fecha_creacion')
    context = {
        'historial_items': historial_items,
        'titulo': 'Historial de Documentos'
    }
    return render(request, 'gestion_escolar/historial.html', context)

@login_required
def historial_detalle_lote(request, historial_id):
    historial_item = get_object_or_404(Historial, id=historial_id)
    if historial_item.lote_reporte:
        lote = historial_item.lote_reporte
        vacancias = lote.vacancias.all()
        context = {
            'historial_item': historial_item,
            'lote': lote,
            'vacancias': vacancias,
            'titulo': f'Detalle de Reporte de Vacancia (Lote {lote.id})'
        }
        return render(request, 'gestion_escolar/historial_detalle_lote.html', context)
    else:
        messages.error(request, "Este registro de historial no está asociado a un lote de vacancia.")
        return redirect('historial')

@login_required
def historial_detalle_tramite(request, historial_id):
    historial_item = get_object_or_404(Historial, id=historial_id)
    if historial_item.datos_tramite:
        # Definimos la lista de claves para la sección "Centro de Trabajo" aquí en la vista.
        ct_keys_list = [
            'techo_financiero_titular', 'clave_ct', 'nombre_ct', 'turno', 
            'domicilio_ct', 'z_escolar', 'region', 'sostenimiento', 
            'supervisor', 'director'
        ]

        context = {
            'historial_item': historial_item,
            'datos_tramite': historial_item.datos_tramite,
            'titulo': f'Detalle de {historial_item.tipo_documento}',
            'ct_keys_list': ct_keys_list, # Pasamos la lista a la plantilla
        }
        return render(request, 'gestion_escolar/historial_detalle_tramite.html', context)
    else:
        messages.error(request, "Este registro de historial no contiene datos de trámite/oficio.")
        return redirect('historial')

@login_required
def descargar_archivo_historial(request, item_id):
    item = get_object_or_404(Historial, id=item_id)
    
    file_path = item.ruta_archivo
    if not file_path and item.lote_reporte and item.lote_reporte.archivo_generado:
        file_path = item.lote_reporte.archivo_generado.path

    if not file_path:
        messages.error(request, "No hay archivo para descargar.")
        return redirect('historial')

    # Security check to prevent directory traversal
    if not os.path.abspath(file_path).startswith(os.path.abspath(settings.BASE_DIR)):
        messages.error(request, "Acceso denegado.")
        return redirect('historial')

    if os.path.exists(file_path):
        try:
            return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=os.path.basename(file_path))
        except Exception as e:
            messages.error(request, f"No se pudo abrir el archivo: {e}")
            return redirect('historial')
    else:
        messages.error(request, "El archivo no fue encontrado en el servidor.")
        return redirect('historial')

@login_required
def eliminar_historial_item(request, item_id):
    if request.method == 'POST':
        item = get_object_or_404(Historial, id=item_id)
        try:
            # Opcional: Eliminar el archivo físico del servidor.
            # Por seguridad, esta línea está comentada.
            # if item.ruta_archivo and os.path.exists(item.ruta_archivo):
            #     os.remove(item.ruta_archivo)
            
            item.delete()
            return JsonResponse({'status': 'success', 'message': 'Registro eliminado correctamente.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

@login_required
def guardar_observacion_historial(request, item_id):
    if request.method == 'POST':
        item = get_object_or_404(Historial, id=item_id)
        try:
            data = json.loads(request.body)
            item.observaciones = data.get('observaciones', '')
            item.save()
            return JsonResponse({'status': 'success', 'message': 'Observación guardada.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f"¡Bienvenido, {username}!")
                return redirect('index') # Redirect to dashboard or a 'next' parameter
            else:
                messages.error(request, "Nombre de usuario o contraseña incorrectos.")
        else:
            messages.error(request, "Por favor, corrige los errores en el formulario.")
    else:
        form = AuthenticationForm()
    
    context = {
        'form': form,
        'titulo': 'Iniciar Sesión'
    }
    return render(request, 'gestion_escolar/login.html', context)

def logout_view(request):
    logout(request)
    messages.info(request, "Has cerrado sesión correctamente.")
    return redirect('login') # Redirect to login page

def signup_view(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'¡Cuenta creada para {username}! Tu solicitud ha sido enviada para aprobación.')
            return redirect('login')
        else:
            messages.error(request, "Por favor, corrige los errores en el formulario.")
    else:
        form = SignUpForm()
    context = {
        'form': form,
        'titulo': 'Solicitud de Registro'
    }
    return render(request, 'gestion_escolar/signup.html', context)

@login_required
def export_maestro_excel(request, pk):
    maestro = get_object_or_404(Maestro, id_maestro=pk)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Detalle_{maestro.id_maestro}"

    # Define headers and data
    data = {
        "ID Maestro": maestro.id_maestro,
        "Nombre Completo": f'{maestro.nombres} {maestro.a_paterno} {maestro.a_materno}',
        "CURP": maestro.curp,
        "RFC": maestro.rfc,
        "Sexo": maestro.get_sexo_display(),
        "Estado Civil": maestro.get_est_civil_display(),
        "Fecha de Nacimiento": maestro.fecha_nacimiento,
        "Techo Financiero": maestro.techo_f,
        "C.C.T.": maestro.id_escuela.id_escuela if maestro.id_escuela else 'N/A',
        "Nombre del Centro de Trabajo": maestro.id_escuela.nombre_ct if maestro.id_escuela else 'N/A',
        "Zona Escolar": maestro.id_escuela.zona_esc.numero if maestro.id_escuela and maestro.id_escuela.zona_esc else 'N/A',
        "Clave Presupuestal": maestro.clave_presupuestal,
        "Categoría": str(maestro.categog) if maestro.categog else '',
        "Código": maestro.codigo,
        "Fecha de Ingreso": maestro.fecha_ingreso,
        "Fecha de Promoción": maestro.fecha_promocion,
        "Formación Académica": maestro.form_academica,
        "Horario": maestro.horario,
        "Función": maestro.get_funcion_display(),
        "Nivel de Estudio": maestro.get_nivel_estudio_display(),
        "Domicilio Particular": maestro.domicilio_part,
        "Población": maestro.poblacion,
        "Código Postal": maestro.codigo_postal,
        "Teléfono": maestro.telefono,
        "Email": maestro.email,
        "Status": maestro.get_status_display(),
        "Observaciones": maestro.observaciones,
    }

    # Write data in two columns: Header and Value
    for key, value in data.items():
        # Handle date objects specifically for formatting
        if isinstance(value, date):
            value = value.strftime("%d/%m/%Y")
        ws.append([key, value])

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    # Prepare the response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="detalle_{maestro.a_paterno}_{maestro.id_maestro}.xlsx"'},
    )
    wb.save(response)

    return response

@login_required
def exportar_maestros_excel(request):
    filtro = request.GET.get('filtro', '')
    funcion = request.GET.get('funcion', '')

    # Construir la consulta base con select_related para optimizar
    maestros_qs = Maestro.objects.select_related('id_escuela', 'id_escuela__zona_esc', 'categog').all().order_by('a_paterno', 'a_materno', 'nombres')

    # Aplicar filtro de función si existe
    if funcion:
        funcion_mapping = {
            'DIRECTOR': {'display': 'Director', 'values': ['DIRECTOR', 'DIRECTOR (A)']},
            'SUPERVISOR': {'display': 'Supervisor', 'values': ['SUPERVISOR', 'SUPERVISOR (A)', 'SUPERVISOR(A)']},
            'MAESTRO_GRUPO': {'display': 'Maestro de Grupo', 'values': ['MAESTRO_GRUPO', 'MAESTRO(A) DE GRUPO', 'MAESTRO(A) DE GRUPO CON ESPECIALIDAD', 'MAESTRO(A) DE GRUPO ESPECIALISTA','MATRO(A) DE GRUPO ESPECIALISTA']},
            'DOCENTE_APOYO': {'display': 'Docente de Apoyo', 'values': ['MAESTRO(A) DE APOYO']},
            'PSICOLOGO': {'display': 'Psicólogo', 'values': ['PSICOLOGO', 'PSICÓLOGO(A)', 'PSICÓLOGO (A)']},
            'TRABAJADOR_SOCIAL': {'display': 'Trabajador Social', 'values': ['TRABAJADOR_SOCIAL', 'TRABAJADOR (A) SOCIAL']},
            'NIÑERO': {'display': 'Niñero', 'values': ['NIÑERO', 'NIÑERO(A)']},
            'SECRETARIO': {'display': 'Secretario', 'values': ['SECRETARIO', 'SECRETARIA ', 'SECRETARIO(A)']},
            'INTENDENTE': {'display': 'Intendente', 'values': ['INTENDENTE']},
            'VELADOR': {'display': 'Velador', 'values': ['VELADOR']},
            'VIGILANTE': {'display': 'Vigilante', 'values': ['VIGILANTE', 'VIGILANTE ']},
            'OTRO': {'display': 'Otro', 'values': ['OTRO']},
            'APOYO_TECNICO_PEDAGOGICO': {'display': 'Apoyo Técnico Pedagógico', 'values': ['APOYO TECNICO PEDAGOGICO']},
            'INSTRUCTOR_TALLER': {'display': 'Instructor de Taller', 'values': ['INSTRUCTOR(A) DE TALLER']},
            'MAESTRO_TALLER': {'display': 'Maestro de Taller', 'values': ['MAESTRO(A) DE TALLER', 'MAESTRO DE TALLER']},
            'MAESTRO_MUSICA': {'display': 'Maestro de Música', 'values': ['MAESTRO(A) MUSICA']},
            'MAESTRO_EDUCACION_FISICA': {'display': 'Maestro de Educación Física', 'values': ['MAESTRO(A) DE EDUCACIÓN FÍSICA']},
            'MEDICO': {'display': 'Médico', 'values': ['MÉDICO(A)', 'MÉDICO (A)']},
            'PROMOTOR_TIC': {'display': 'Promotor TIC', 'values': ['PROMOTOR TIC', 'PROMOTOR TIC ']},
            'TERAPISTA_FISICO': {'display': 'Terapista Físico', 'values': ['TERAPISTA FISICO ']},
            'BIBLIOTECARIO': {'display': 'Bibliotecario', 'values': ['BIBLIOTECARIO(A)']},
            'ADMINISTRATIVO_ESPECIALIZADO': {'display': 'Administrativo Especializado', 'values': ['ADMINISTRATIVO ESPECIALIZADO']},
            'OFICIAL_SERVICIOS_MANTENIMIENTO': {'display': 'Oficial de Servicios y Mantenimiento', 'values': ['OFICIAL DE SERVICIOS Y MANTENIMIENTO', 'OFICIAL DE SERVICIOS DE MANTENIMIENTO']},
            'ASISTENTE_DE_SERVICIOS': {'display': 'Asistente de Servicios', 'values': ['ASISTENTE DE SERVICIOS']},
            'ASESOR_JURIDICO': {'display': 'Asesor Jurídico', 'values': ['ASESOR JURÍDICO']},
            'AUXILIAR_DE_GRUPO': {'display': 'Auxiliar de Grupo', 'values': ['AUXILIAR DE GRUPO']},
            'MAESTRO_COMUNICACION': {'display': 'Maestro de Comunicación', 'values': ['MAESTRO(A) DE COMUNICACIÓN ']},
            'MAESTRO_AULA_HOSPITALARIA': {'display': 'Maestro Aula Hospitalaria', 'values': ['MAESTRO(A) AULA HOSPITALARIA']},
        }
        funcion_info = funcion_mapping.get(funcion)
        if funcion_info:
            maestros_qs = maestros_qs.filter(funcion__in=funcion_info['values'])

    # Aplicar filtro de texto libre si existe
    if filtro:
        maestros_qs = maestros_qs.filter(
            Q(nombres__icontains=filtro) |
            Q(a_paterno__icontains=filtro) |
            Q(a_materno__icontains=filtro) |
            Q(rfc__icontains=filtro) |
            Q(curp__icontains=filtro) |
            Q(id_maestro__icontains=filtro) |
            Q(id_escuela__nombre_ct__icontains=filtro) |
            Q(id_escuela__id_escuela__icontains=filtro) |
            Q(categog__descripcion__icontains=filtro)
        )

    # Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maestros"

    # Definir los encabezados
    headers = [
        "ID Maestro", "Nombre(s)", "Apellido Paterno", "Apellido Materno", "RFC", "CURP",
        "Sexo", "Estado Civil", "Fecha Nacimiento", "Techo Financiero",
        "CCT", "Nombre del CT", "Zona Escolar", "Función", "Categoría",
        "Clave Presupuestal", "Código", "Fecha Ingreso", "Fecha Promoción",
        "Formación Académica", "Horario", "Nivel de Estudio", "Domicilio Particular",
        "Población", "Código Postal", "Teléfono", "Email", "Status", "Observaciones"
    ]
    ws.append(headers)

    # Escribir los datos de cada maestro
    for maestro in maestros_qs:
        escuela = maestro.id_escuela
        zona_numero = ''
        if escuela and escuela.zona_esc:
            zona_numero = escuela.zona_esc.numero

        row = [
            maestro.id_maestro,
            maestro.nombres,
            maestro.a_paterno,
            maestro.a_materno,
            maestro.rfc,
            maestro.curp,
            maestro.get_sexo_display(),
            maestro.get_est_civil_display(),
            maestro.fecha_nacimiento.strftime("%Y-%m-%d") if maestro.fecha_nacimiento else '',
            maestro.techo_f,
            escuela.id_escuela if escuela else '',
            escuela.nombre_ct if escuela else '',
            zona_numero,
            maestro.get_funcion_display(),
            maestro.categog.descripcion if maestro.categog else '',
            maestro.clave_presupuestal,
            maestro.codigo,
            maestro.fecha_ingreso.strftime("%Y-%m-%d") if maestro.fecha_ingreso else '',
            maestro.fecha_promocion.strftime("%Y-%m-%d") if maestro.fecha_promocion else '',
            maestro.form_academica,
            maestro.horario,
            maestro.get_nivel_estudio_display(),
            maestro.domicilio_part,
            maestro.poblacion,
            maestro.codigo_postal,
            maestro.telefono,
            maestro.email,
            maestro.get_status_display(),
            maestro.observaciones,
        ]
        ws.append(row)

    # Preparar la respuesta para la descarga
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="reporte_maestros.xlsx"'},
    )
    wb.save(response)

    return response

@permission_required('gestion_escolar.acceder_vacancias', raise_exception=True)
def gestionar_lote_vacancia(request):
    # Busca un lote en proceso para el usuario actual o crea uno nuevo
    # Primero, intenta obtener un lote existente
    lotes_en_proceso = LoteReporteVacancia.objects.filter(
        usuario_generador=request.user,
        estado='EN_PROCESO'
    ).order_by('-fecha_creacion')

    if lotes_en_proceso.exists():
        lote = lotes_en_proceso.first()
        # Si hay más de uno, marcar los demás como cancelados o eliminarlos
        if lotes_en_proceso.count() > 1:
            for old_lote in lotes_en_proceso[1:]:
                old_lote.estado = 'CANCELADO' # O eliminar: old_lote.delete()
                old_lote.save()
    else:
        lote = LoteReporteVacancia.objects.create(
            usuario_generador=request.user,
            estado='EN_PROCESO'
        )

    if request.method == 'POST':
        form = VacanciaForm(request.POST)
        if form.is_valid():
            # El formulario es válido, pero aún no guardamos la vacancia.
            # Primero, necesitamos calcular todos los campos derivados.
            maestro = form.cleaned_data['maestro_titular']
            escuela = maestro.id_escuela

            # Iniciar la creación de la instancia de Vacancia
            vacancia = form.save(commit=False)
            vacancia.lote = lote

            # Manejar Maestro Interino
            maestro_interino_obj = form.cleaned_data.get('maestro_interino')
            if maestro_interino_obj:
                vacancia.nombre_interino = f'{maestro_interino_obj.nombres} {maestro_interino_obj.a_paterno} {maestro_interino_obj.a_materno}'
                vacancia.curp_interino = maestro_interino_obj.curp

                # Obtener datos de prelación si existen
                if maestro_interino_obj.curp:
                    prelacion = Prelacion.objects.filter(curp=maestro_interino_obj.curp).first()
                    if prelacion:
                        vacancia.posicion_orden = prelacion.pos_orden
                        vacancia.folio_prelacion = prelacion.folio

            # --- Lógica de Transformación (portada de VBA) ---
            # 1. Dirección
            vacancia.direccion = f"{escuela.nombre_ct}, {escuela.domicilio}, DURANGO, {escuela.region}, {escuela.get_turno_display()}, ZONA ECONOMICA:{escuela.zona_economica}"

            # 2. Destino
            apreciacion_desc = form.cleaned_data['apreciacion'].descripcion
            if apreciacion_desc.startswith("ADMISIÓN"):
                vacancia.destino = "Admisión"
            elif apreciacion_desc.startswith("PROMOCIÓN"):
                vacancia.destino = "Promoción vertical"
            else:
                vacancia.destino = ""

            # 3. Sostenimiento y Turno
            vacancia.sostenimiento = "Federalizado" if escuela.sostenimiento == 'FEDERAL' else "Estatal"
            vacancia.turno = escuela.get_turno_display()

            # 4. Tipo de Movimiento para Reporte
            vacancia.tipo_movimiento_reporte = form.cleaned_data['tipo_movimiento_original']

            # 5. Tipo Plaza y Horas
            hrs = maestro.hrs or "00.0"
            vacancia.tipo_plaza = "JORNADA" if hrs == "00.0" else "HORA/SEMANA/MES"
            if hrs == "00.0":
                vacancia.horas = None
            else:
                try:
                    # Convertir a int para eliminar ceros iniciales y luego a str
                    vacancia.horas = str(int(float(hrs)))
                except (ValueError, TypeError):
                    vacancia.horas = None # En caso de que hrs no sea un número válido

            # 6. Otros campos directos del maestro o escuela
            vacancia.municipio = escuela.region
            vacancia.zona_economica = f"Zona {escuela.zona_economica}"
            vacancia.categoria = maestro.categog.id_categoria if maestro.categog else ''
            vacancia.clave_presupuestal = maestro.clave_presupuestal
            vacancia.techo_financiero = maestro.techo_f
            vacancia.clave_ct = escuela.id_escuela
            vacancia.nombre_titular_reporte = f'{maestro.nombres} {maestro.a_paterno} {maestro.a_materno}'

            vacancia.save()

            # --- Crear registro en el historial para la vacancia ---
            maestro_titular_obj = vacancia.maestro_titular
            maestro_interino_obj = vacancia.maestro_interino

            datos_tramite_vacancia = {
                "tipo_movimiento": "Detalle de Asignación de Vacancia",
                "id_vacancia": vacancia.id,
                "clave_presupuestal_posicion": vacancia.clave_presupuestal,
                "maestro_titular_info": {
                    "id_maestro": maestro_titular_obj.id_maestro,
                    "nombre_completo": get_full_name(maestro_titular_obj),
                    "clave_presupuestal": maestro_titular_obj.generar_clave_presupuestal()
                }
            }

            if maestro_interino_obj:
                datos_tramite_vacancia["maestro_interino_info"] = {
                    "id_maestro": maestro_interino_obj.id_maestro,
                    "nombre_completo": get_full_name(maestro_interino_obj),
                    "clave_presupuestal": maestro_interino_obj.generar_clave_presupuestal()
                }

            Historial.objects.create(
                usuario=request.user,
                tipo_documento="Asignación de Vacancia",
                maestro=maestro_titular_obj,
                maestro_secundario_nombre=get_full_name(maestro_interino_obj) if maestro_interino_obj else '',
                ruta_archivo="", # No hay archivo específico para esta entrada de historial de vacancia
                motivo="Asignación de Vacancia",
                lote_reporte=lote,
                datos_tramite=datos_tramite_vacancia
            )
            # --- Fin del registro en historial ---

            messages.success(request, "Vacancia agregada al lote actual.")
            return redirect('gestionar_lote_vacancia')
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = VacanciaForm()

    vacancias_en_lote = Vacancia.objects.filter(lote=lote).order_by('-id')
    context = {
        'form': form,
        'lote': lote,
        'vacancias_en_lote': vacancias_en_lote,
        'titulo': 'Generar Reporte de Vacancia'
    }
    return render(request, 'gestion_escolar/gestionar_lote_vacancia.html', context)

from django.views.decorators.clickjacking import xframe_options_exempt

def _get_lote_y_vacancias(request, lote_id):
    lote = get_object_or_404(LoteReporteVacancia, id=lote_id, usuario_generador=request.user)
    vacancias = lote.vacancias.all()

    if not vacancias.exists():
        raise ValueError("No hay vacancias en este lote para exportar.")

    if lote.estado == 'GENERADO':
        messages.warning(request, "Este lote ya fue procesado anteriormente.")
        return JsonResponse({'status': 'error', 'message': 'Este lote ya fue procesado.'}, status=400)

    lote.estado = 'PROCESANDO'
    lote.save()
    return lote, vacancias

@login_required
@transaction.atomic
def exportar_paso_word(request, lote_id):
    print("******** PASO 1: GENERANDO DOCUMENTOS WORD ********")
    try:
        lote, vacancias = _get_lote_y_vacancias(request, lote_id)
    except ValueError as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    plantilla_solicitud_asignacion = PlantillaTramite.objects.filter(nombre="SOLICITUD DE ASIGNACION").first()
    documentos_word_generados = 0
    word_docs_info = []

    if not plantilla_solicitud_asignacion:
        return JsonResponse({'status': 'warning', 'message': 'Plantilla "SOLICITUD DE ASIGNACION" no encontrada. Saltando paso de Word.', 'word_count': 0, 'word_docs': []})

    # 1. GENERAR DOCUMENTOS WORD (para vacancias ≤ 3 meses con interino)
    for vacancia in vacancias:
        if vacancia.maestro_interino:
            # Verificar si la vacancia es ≤ 3 meses
            if vacancia.fecha_inicio and vacancia.fecha_final:
                duration_months = get_month_diff(vacancia.fecha_inicio, vacancia.fecha_final)
                if duration_months <= 3:
                    print(f"DEBUG: Generando Word para vacancia ID {vacancia.id} (≤ 3 meses)")
                    
                    form_data_for_word = {
                        'plantilla': plantilla_solicitud_asignacion,
                        'maestro_titular': vacancia.maestro_titular,
                        'maestro_interino': vacancia.maestro_interino,
                        'fecha_efecto1': vacancia.fecha_inicio, 
                        'fecha_efecto2': vacancia.fecha_final,
                        'fecha_efecto3': vacancia.fecha_inicio, 
                        'fecha_efecto4': vacancia.fecha_final,
                        'folio': vacancia.folio_prelacion, 
                        'observaciones': vacancia.observaciones,
                        'no_prel_display': vacancia.posicion_orden, 
                        'folio_prel_display': vacancia.folio_prelacion,
                    }
                    
                    motivo_tramite_obj = MotivoTramite.objects.filter(motivo_tramite=vacancia.tipo_movimiento_original).first()
                    form_data_for_word['motivo_tramite'] = motivo_tramite_obj
                    
                    tipo_val_display = ''
                    if vacancia.maestro_interino.curp:
                        prelacion = Prelacion.objects.filter(curp=vacancia.maestro_interino.curp).first()
                        if prelacion:
                            tipo_val_display = prelacion.tipo_val
                    form_data_for_word['tipo_val_display'] = tipo_val_display

                    success, doc_path = generate_word_document(form_data_for_word, plantilla_solicitud_asignacion, request.user)
                    if success:
                        try:
                            historial_word = Historial.objects.create(
                                usuario=request.user,
                                tipo_documento=f"Oficio - {plantilla_solicitud_asignacion.nombre}",
                                maestro=vacancia.maestro_titular,
                                ruta_archivo=doc_path,
                                motivo=motivo_tramite_obj.motivo_tramite if motivo_tramite_obj else '',
                                maestro_secundario_nombre=get_full_name(vacancia.maestro_interino),
                                datos_tramite=serialize_form_data(form_data_for_word)
                            )
                            word_docs_info.append({
                                'id': historial_word.id,
                                'nombre': os.path.basename(doc_path),
                                'url': reverse('descargar_archivo_historial', args=[historial_word.id])
                            })
                            documentos_word_generados += 1
                        except Exception as e:
                            print(f"DEBUG: ❌ Error creando historial para Word: {e}")
    
    return JsonResponse({
        'status': 'success',
        'message': f'Se generaron {documentos_word_generados} documento(s) Word.',
        'word_count': documentos_word_generados,
        'word_docs': word_docs_info
    })

@login_required
@transaction.atomic
def exportar_paso_gsheets(request, lote_id):
    print("******** PASO 2: ENVIANDO A GOOGLE SHEETS ********")
    try:
        lote, vacancias = _get_lote_y_vacancias(request, lote_id)
    except ValueError as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    vacancias_enviadas = 0
    errores_gsheets = []
    for vacancia in vacancias:
        if vacancia.maestro_interino and vacancia.fecha_inicio and vacancia.fecha_final:
            duration_months = get_month_diff(vacancia.fecha_inicio, vacancia.fecha_final)
            if duration_months <= 3:
                print(f"DEBUG: Enviando a Google Sheets: {get_full_name(vacancia.maestro_interino)}")
                # Construir la fila de datos para Google Sheets
                google_sheet_row_data = [
                    str(vacancias.filter(id__lte=vacancia.id).count()),
                    datetime.now().strftime("%Y-%m-%d"),
                    "EDUCACIÓN ESPECIAL",
                    "Durango",
                    vacancia.municipio or '',
                    vacancia.direccion or '',
                    vacancia.region or '',
                    vacancia.zona_economica or '',
                    vacancia.destino or '',
                    vacancia.apreciacion.descripcion if vacancia.apreciacion else '',
                    vacancia.get_tipo_vacante_display() or '',
                    vacancia.tipo_plaza or '',
                    vacancia.horas if vacancia.tipo_plaza == "HORA/SEMANA/MES" else '',
                    vacancia.sostenimiento or '',
                    vacancia.fecha_inicio.strftime("%Y-%m-%d") if vacancia.fecha_inicio else '',
                    vacancia.fecha_final.strftime("%Y-%m-%d") if vacancia.fecha_final else '',
                    vacancia.categoria or '',
                    vacancia.pseudoplaza or '',
                    vacancia.clave_presupuestal or '',
                    vacancia.techo_financiero or '',
                    vacancia.clave_ct or '',
                    vacancia.turno or '',
                    vacancia.tipo_movimiento_original or '',
                    vacancia.observaciones or '',
                    vacancia.maestro_interino.curp or '',
                    '',
                    vacancia.maestro_interino.form_academica or '',
                    "N/A" if vacancia.tipo_plaza == "JORNADA" else (vacancia.apreciacion.descripcion if vacancia.apreciacion else ''),
                    '', '', '', '',
                    format_date_for_solicitud_asignacion(vacancia.fecha_inicio) if vacancia.fecha_inicio else '',
                    format_date_for_solicitud_asignacion(vacancia.fecha_final) if vacancia.fecha_final else '',
                    '', '',
                    f"DEE/{vacancia.folio_prelacion}/2025" if vacancia.folio_prelacion else '',
                    '', '',
                ]
                success_gs, message_gs = send_to_google_sheet(google_sheet_row_data)
                if not success_gs:
                    error_msg = f"Fallo al enviar datos del interino '{get_full_name(vacancia.maestro_interino)}': {message_gs}"
                    print(f"DEBUG: ❌ Google Sheets - {error_msg}")
                    errores_gsheets.append(error_msg)
                else:
                    vacancias_enviadas += 1
    
    mensaje_final = f'Se enviaron datos de {vacancias_enviadas} vacancia(s) a Google Sheets.'
    if errores_gsheets:
        mensaje_final += f' Hubo {len(errores_gsheets)} error(es).'

    return JsonResponse({
        'status': 'success',
        'message': mensaje_final,
        'gsheets_count': vacancias_enviadas,
        'gsheets_errors': errores_gsheets
    })

@login_required
@transaction.atomic
def exportar_paso_excel(request, lote_id):
    print("******** PASO 3: GENERANDO ARCHIVO EXCEL ********")
    try:
        lote, vacancias = _get_lote_y_vacancias(request, lote_id)
        print("DEBUG: Iniciando generación de Excel...")
        
        # Verificar que el directorio de templates existe
        template_path = os.path.join(settings.BASE_DIR, 'tramites', 'Plantillas', 'Excel', 'FORMATOVACANCIAUSICAMM.xlsx')
        print(f"DEBUG: Buscando template en: {template_path}")
        
        if not os.path.exists(template_path):
            # Intentar rutas alternativas
            alternative_paths = [
                os.path.join(settings.BASE_DIR, 'gestion_escolar', 'templates', 'tramites', 'Plantillas', 'Excel', 'FORMATOVACANCIAUSICAMM.xlsx'),
                os.path.join(settings.BASE_DIR, 'templates', 'tramites', 'Plantillas', 'Excel', 'FORMATOVACANCIAUSICAMM.xlsx'),
            ]
            
            template_found = False
            for alt_path in alternative_paths:
                if os.path.exists(alt_path):
                    template_path = alt_path
                    template_found = True
                    print(f"DEBUG: Template encontrado en ruta alternativa: {template_path}")
                    break
            
            if not template_found:
                raise FileNotFoundError(f"No se encontró el template en ninguna ruta: {template_path}")

        print("DEBUG: Cargando workbook...")
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active
        
        print("DEBUG: Preparando datos para Excel...")
        row_num = 2
        campos_excel = [
            'nivel', 'entidad', 'municipio', 'direccion', 'region', 'zona_economica', 'destino', 'apreciacion',
            'tipo_vacante', 'tipo_plaza', 'horas', 'sostenimiento', 'fecha_inicio', 'fecha_final', 'categoria',
            'pseudoplaza', 'clave_presupuestal', 'techo_financiero', 'clave_ct', 'turno', 'tipo_movimiento_reporte', 'observaciones',
            'posicion_orden', 'folio_prelacion', 'curp_interino', 'nombre_interino'
        ]
        
        for vacancia in vacancias:
            print(f"DEBUG: Procesando vacancia {vacancia.id} para Excel...")
            for i, field_name in enumerate(campos_excel):
                cell = sheet.cell(row=row_num, column=i + 1)
                valor = ''
                
                try:
                    if field_name == 'curp_interino': 
                        valor = vacancia.maestro_interino.curp if vacancia.maestro_interino else vacancia.curp_interino or ''
                    elif field_name == 'nombre_interino': 
                        valor = get_full_name(vacancia.maestro_interino) if vacancia.maestro_interino else vacancia.nombre_interino or ''
                    elif field_name == 'apreciacion': 
                        valor = vacancia.apreciacion.descripcion if vacancia.apreciacion else ''
                    elif field_name == 'tipo_vacante': 
                        valor = vacancia.get_tipo_vacante_display().capitalize() if vacancia.tipo_vacante else ''
                    elif field_name == 'zona_economica' and isinstance(valor, str):
                        valor = {'Zona II': 'Zona 2', 'Zona III': 'Zona 3'}.get(valor, valor)
                    else: 
                        valor = getattr(vacancia, field_name, '') or ''
                        
                    # Convertir fechas a string
                    if field_name in ['fecha_inicio', 'fecha_final'] and valor:
                        if isinstance(valor, (date, datetime)):
                            valor = valor.strftime("%Y-%m-%d")
                            
                except Exception as field_error:
                    print(f"DEBUG: Error procesando campo {field_name} para vacancia {vacancia.id}: {field_error}")
                    valor = f"Error: {field_error}"
                
                cell.value = valor
                
            row_num += 1

        print("DEBUG: Guardando archivo Excel...")
        output_dir = os.path.join(settings.MEDIA_ROOT, 'reportes_vacancias')
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"VACANCIA_{timestamp}.xlsx"
        output_path_server = os.path.join(output_dir, output_filename)
        
        workbook.save(output_path_server)
        print(f"DEBUG: Excel guardado en: {output_path_server}")

        # Crear registro en historial
        historial_excel = Historial.objects.create(
            usuario=request.user, 
            tipo_documento="Reporte de Vacancia", 
            maestro=None,
            ruta_archivo=output_path_server, 
            motivo="Reporte de Vacancia", 
            lote_reporte=lote
        )
        
        # Actualizar lote
        lote.archivo_generado = os.path.join('reportes_vacancias', output_filename)
        lote.estado = 'GENERADO'
        lote.fecha_generado = datetime.now()
        lote.save()

        # Preparar respuesta - SIMPLIFICADA PARA EVITAR ERRORES
        response_data = {
            'status': 'success',
            'message': f'Lote procesado exitosamente. Se generaron {documentos_word_generados} documento(s) Word y 1 archivo Excel.',
            'excel_url': reverse('descargar_archivo_historial', args=[historial_excel.id]),
            'excel_name': output_filename
        }

        print("DEBUG: Proceso completado exitosamente, retornando respuesta JSON")
        return JsonResponse(response_data)

    except Exception as e:
        print(f"DEBUG: ❌ Error generando Excel: {str(e)}")
        import traceback
        print(f"DEBUG: Traceback completo: {traceback.format_exc()}")
        
        messages.error(request, f"Error al generar el archivo Excel: {str(e)}")
        lote.estado = 'EN_PROCESO'
        lote.save()
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def get_maestro_data_for_vacancia(request):
    maestro_id = request.GET.get('maestro_id')
    data = {}
    if maestro_id:
        try:
            maestro = Maestro.objects.get(id_maestro=maestro_id)
            data = {
                'nombre_completo': f'{maestro.nombres} {maestro.a_paterno} {maestro.a_materno}',
                'clave_presupuestal': maestro.clave_presupuestal,
                'categoria': maestro.categog.id_categoria if maestro.categog else '',
                'curp': maestro.curp or ''
            }
        except Maestro.DoesNotExist:
            data = {'error': 'Maestro no encontrado'}
    return JsonResponse(data)

@login_required
def get_interino_and_prelacion_data_ajax(request):
    maestro_id = request.GET.get('maestro_id')
    data = {
        'curp_interino': '',
        'folio_prelacion': '',
        'posicion_orden': '',
        'tipo_val': '',
        'error': ''
    }

    if maestro_id:
        try:
            maestro = Maestro.objects.get(id_maestro=maestro_id)
            data['curp_interino'] = maestro.curp or '' # Siempre devolver la CURP si el maestro existe

            # Buscar en la tabla Prelacion usando la CURP del interino
            if maestro.curp:
                prelacion = Prelacion.objects.filter(curp=maestro.curp).first()
                if prelacion:
                    data['folio_prelacion'] = prelacion.folio or ''
                    data['posicion_orden'] = prelacion.pos_orden or ''
                    data['tipo_val'] = prelacion.tipo_val or ''
                # else: No establecer data['error'] aquí, solo significa que no hay datos de prelación
            # else: No establecer data['error'] aquí, solo significa que el maestro no tiene CURP registrada, pero el maestro sí existe

        except Maestro.DoesNotExist:
            data['error'] = 'Maestro interino no encontrado.'
            data['curp_interino'] = 'N/A' # Asegurarse de que la CURP también refleje el error
        except Exception as e:
            data['error'] = f'Error inesperado: {str(e)}'
            data['curp_interino'] = 'Error' # Asegurarse de que la CURP también refleje el error

    return JsonResponse(data)

@login_required
def eliminar_vacancia_lote(request, pk):
    if request.method == 'POST':
        vacancia = get_object_or_404(Vacancia, pk=pk)
        # Asegurarse de que la vacancia pertenece al lote en proceso del usuario actual
        if vacancia.lote.usuario_generador == request.user and vacancia.lote.estado == 'EN_PROCESO':
            vacancia.delete()
            messages.success(request, "Vacancia eliminada del lote correctamente.")
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error', 'message': 'No autorizado para eliminar esta vacancia.'}, status=403)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

# --- VISTAS PARA PENDIENTES ---

class PendienteCreateView(LoginRequiredMixin, CreateView):
    form_class = PendienteForm
    template_name = 'gestion_escolar/pendiente_form.html'
    success_url = reverse_lazy('pendientes_activos')

    def form_valid(self, form):
        form.instance.usuario = self.request.user
        messages.success(self.request, "Pendiente creado correctamente.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Nuevo Pendiente'
        return context

class PendienteActiveListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'gestion_escolar.acceder_pendientes'
    model = Pendiente
    template_name = 'gestion_escolar/pendiente_list.html'
    context_object_name = 'pendientes'

    def get_queryset(self):
        today = timezone.now().date()
        return Pendiente.objects.filter(
            usuario=self.request.user,
            completado=False,
            fecha_programada__lte=today
        ).order_by('fecha_programada')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Mis Pendientes Activos'
        context['vista_todos'] = False
        return context

class PendienteAllListView(LoginRequiredMixin, ListView):
    model = Pendiente
    template_name = 'gestion_escolar/pendiente_list.html'
    context_object_name = 'pendientes'

    def get_queryset(self):
        return Pendiente.objects.filter(usuario=self.request.user).order_by('-fecha_programada')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Todos Mis Pendientes'
        context['vista_todos'] = True
        return context

@login_required
def pendiente_marcar_completado(request, pk):
    if request.method == 'POST':
        pendiente = get_object_or_404(Pendiente, pk=pk, usuario=request.user)
        pendiente.completado = True
        pendiente.save()
        messages.success(request, f"El pendiente '{pendiente.titulo}' ha sido marcado como completado.")
        return redirect('pendientes_activos')
    else:
        return redirect('pendientes_activos')

# --- VISTAS PARA CORRESPONDENCIA ---

class CorrespondenciaInboxView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    permission_required = 'gestion_escolar.acceder_bandeja_entrada'
    model = Correspondencia
    template_name = 'gestion_escolar/correspondencia_inbox.html'
    context_object_name = 'mensajes'

    def get_queryset(self):
        return Correspondencia.objects.filter(destinatario=self.request.user).order_by('-fecha_creacion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Bandeja de Entrada'
        return context

class CorrespondenciaDetailView(LoginRequiredMixin, DetailView):
    model = Correspondencia
    template_name = 'gestion_escolar/correspondencia_detail.html'
    context_object_name = 'mensaje'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if obj.destinatario == self.request.user:
            if not obj.leido:
                obj.leido = True
                obj.save()
            return obj
        else:
            raise PermissionDenied("No tienes permiso para ver este mensaje.")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = self.object.asunto
        return context

@login_required
def correspondencia_eliminar(request, pk):
    if request.method == 'POST':
        mensaje = get_object_or_404(Correspondencia, pk=pk)
        if mensaje.destinatario == request.user:
            mensaje.delete()
            messages.success(request, "Mensaje eliminado correctamente.")
        else:
            messages.error(request, "No tienes permiso para eliminar este mensaje.")
    return redirect('correspondencia_inbox')


class CorrespondenciaCreateView(LoginRequiredMixin, CreateView):


    form_class = CorrespondenciaForm


    template_name = 'gestion_escolar/correspondencia_form.html'


    success_url = reverse_lazy('correspondencia_inbox')





    def form_valid(self, form):


        form.instance.remitente = self.request.user


        messages.success(self.request, "Mensaje enviado correctamente.")


        return super().form_valid(form)





    def get_form(self, form_class=None):


        form = super().get_form(form_class)


        # Excluir al usuario actual de la lista de destinatarios


        form.fields['destinatario'].queryset = User.objects.exclude(pk=self.request.user.pk)


        return form





        def get_context_data(self, **kwargs):
            context = super().get_context_data(**kwargs)
            context['titulo'] = 'Redactar Nuevo Mensaje'
            return context



class RegistroCorrespondenciaListView(LoginRequiredMixin, ListView):
    model = RegistroCorrespondencia
    template_name = 'gestion_escolar/registrocorrespondencia_list.html'
    context_object_name = 'registros'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Registro de Correspondencia'
        return context

class RegistroCorrespondenciaCreateView(LoginRequiredMixin, CreateView):
    model = RegistroCorrespondencia
    form_class = RegistroCorrespondenciaForm
    template_name = 'gestion_escolar/registrocorrespondencia_form.html'
    success_url = reverse_lazy('registrocorrespondencia_list')

class RegistroCorrespondenciaDetailView(LoginRequiredMixin, DetailView):
    model = RegistroCorrespondencia
    template_name = 'gestion_escolar/registrocorrespondencia_detail.html'
    context_object_name = 'registro'

class RegistroCorrespondenciaUpdateView(LoginRequiredMixin, UpdateView):
    model = RegistroCorrespondencia
    form_class = RegistroCorrespondenciaForm
    template_name = 'gestion_escolar/registrocorrespondencia_form.html'
    success_url = reverse_lazy('registrocorrespondencia_list')

class RegistroCorrespondenciaDeleteView(LoginRequiredMixin, DeleteView):
    model = RegistroCorrespondencia
    template_name = 'gestion_escolar/registrocorrespondencia_confirm_delete.html'
    success_url = reverse_lazy('registrocorrespondencia_list')

class RoleListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Group
    template_name = 'gestion_escolar/role_list.html'
    context_object_name = 'roles'

    def test_func(self):
        return self.request.user.is_superuser

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Gestión de Roles y Permisos'
        context['users'] = User.objects.all()
        return context

def get_permissions_matrix():
    from django.contrib.contenttypes.models import ContentType
    from django.contrib.auth.models import Permission

    matrix = []
    # Define los modelos que quieres en la matriz y su orden
    ordered_models = [
        'zona', 'escuela', 'maestro', 'categoria', 
        'historial', 'pendiente', 'correspondencia', 'registrocorrespondencia'
    ]
    
    app_models = ContentType.objects.filter(app_label='gestion_escolar').order_by('model')

    for model_name in ordered_models:
        try:
            ct = app_models.get(model=model_name)
            perms = Permission.objects.filter(content_type=ct)
            matrix.append({
                'model_name': ct.name,
                'type': 'crud',
                'view': perms.filter(codename__startswith='view_').first(),
                'add': perms.filter(codename__startswith='add_').first(),
                'change': perms.filter(codename__startswith='change_').first(),
                'delete': perms.filter(codename__startswith='delete_').first(),
            })
        except ContentType.DoesNotExist:
            continue

    # Añadir permisos personalizados
    custom_perms_codenames = [
        'acceder_oficios', 'acceder_tramites', 'acceder_vacancias', 
        'acceder_historial', 'acceder_ajustes', 'acceder_bandeja_entrada',
        'acceder_reportes', 'acceder_pendientes',
        # New dashboard permissions
        'ver_estadisticas_generales', 'ver_grafico_distribucion_zona',
        'ver_lista_pendientes', 'ver_lista_ultimo_personal', 'ver_ultima_correspondencia',
        'acceder_kardex',
    ]
    
    for codename in custom_perms_codenames:
        try:
            perm = Permission.objects.get(codename=codename)
            matrix.append({
                'model_name': perm.name,
                'type': 'custom',
                'permission': perm
            })
        except Permission.DoesNotExist:
            continue
            
    return matrix

class RoleCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Group
    form_class = RolePermissionForm
    template_name = 'gestion_escolar/role_form.html'
    success_url = reverse_lazy('role_list')

    def test_func(self):
        return self.request.user.is_superuser

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Crear Nuevo Rol'
        context['permissions_matrix'] = get_permissions_matrix()
        return context

class RoleUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Group
    form_class = RolePermissionForm
    template_name = 'gestion_escolar/role_form.html'
    success_url = reverse_lazy('role_list')

    def test_func(self):
        return self.request.user.is_superuser

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = f'Editar Rol: {self.object.name}'
        context['permissions_matrix'] = get_permissions_matrix()
        return context

@login_required
@user_passes_test(lambda u: u.is_superuser)
def manage_role_members(request, pk):
    role = get_object_or_404(Group, pk=pk)
    users_in_role = role.user_set.all()
    users_not_in_role = User.objects.exclude(groups=role)

    if request.method == 'POST':
        users_to_add = request.POST.getlist('users_to_add')
        users_to_remove = request.POST.getlist('users_to_remove')

        for user_id in users_to_add:
            user = User.objects.get(pk=user_id)
            user.groups.add(role)
        
        for user_id in users_to_remove:
            user = User.objects.get(pk=user_id)
            user.groups.remove(role)
        
        messages.success(request, 'Miembros del rol actualizados correctamente.')
        return redirect('role_members', pk=pk)

    context = {
        'titulo': f'Gestionar Miembros del Rol: {role.name}',
        'role': role,
        'members': users_in_role,
        'non_members': users_not_in_role
    }
    return render(request, 'gestion_escolar/role_members.html', context)

class RoleDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Group
    template_name = 'gestion_escolar/role_confirm_delete.html'
    success_url = reverse_lazy('role_list')
    def test_func(self):
        return self.request.user.is_superuser

@login_required
def reporte_distribucion_funcion(request):
    maestros_qs = Maestro.objects.all()

    # Get filter parameters
    zona_id = request.GET.get('zona')
    escuela_id = request.GET.get('escuela')

    if zona_id:
        maestros_qs = maestros_qs.filter(id_escuela__zona_esc_id=zona_id)
    
    if escuela_id:
        maestros_qs = maestros_qs.filter(id_escuela_id=escuela_id)

    # Group by function and count
    distribucion = maestros_qs.values('funcion').annotate(total=Count('funcion')).order_by('-total')

    # Prepare data for Chart.js
    labels = [d['funcion'] for d in distribucion]
    data = [d['total'] for d in distribucion]

    # Get all zonas and escuelas for the filters
    zonas = Zona.objects.all()
    escuelas = Escuela.objects.all()

    context = {
        'titulo': 'Distribución de Personal por Función',
        'labels': json.dumps(labels),
        'data': json.dumps(data),
        'zonas': zonas,
        'escuelas': escuelas,
        'selected_zona': int(zona_id) if zona_id else None,
        'selected_escuela': int(escuela_id) if escuela_id else None,
    }
    return render(request, 'gestion_escolar/reporte_distribucion_funcion.html', context)

# --- VISTAS PARA KARDEX ---

@login_required
def kardex_maestros_ajax(request):
    draw = int(request.GET.get('draw', 0))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')

    order_column_index = int(request.GET.get('order[0][column]', 0))
    order_dir = request.GET.get('order[0][dir]', 'asc')
    column_names = ['a_paterno', 'clave_presupuestal'] 
    order_column = column_names[order_column_index]
    if order_dir == 'desc':
        order_column = f'-{order_column}'

    queryset = Maestro.objects.all().exclude(id_maestro__isnull=True).exclude(id_maestro='')
    records_total = queryset.count()

    if search_value:
        from unidecode import unidecode
        search_value_upper = search_value.upper() # Convert search value to uppercase
        search_terms = search_value_upper.split() # Use uppercase search value
        table_name = Maestro._meta.db_table
        
        where_clauses = []
        params = []

        for term in search_terms:
            unaccented_term = f'%{unidecode(term)}%'
            term_with_wildcards = f'%{term}%'
            
            term_clause = f"""(
                UPPER(unaccent({table_name}.nombres)) LIKE %s OR
                UPPER(unaccent({table_name}.a_paterno)) LIKE %s OR
                UPPER(unaccent({table_name}.a_materno)) LIKE %s OR
                UPPER({table_name}.clave_presupuestal) LIKE %s
            )"""
            where_clauses.append(term_clause)
            
            params.extend([unaccented_term, unaccented_term, unaccented_term, term_with_wildcards])

        if where_clauses:
            queryset = queryset.extra(where=[" AND ".join(where_clauses)], params=params)

    records_filtered = queryset.count()
    queryset = queryset.order_by(order_column)[start:start + length]

    data = []
    for maestro in queryset:
        kardex_url = reverse("kardex_maestro_detail", args=[maestro.pk]) + "?from=lista"
        actions = f'<a href="{kardex_url}" class="btn btn-sm btn-warning">Ver Kardex</a>'
        data.append([
            f'{maestro.a_paterno} {maestro.a_materno} {maestro.nombres}',
            maestro.clave_presupuestal or '-',
            actions
        ])

    response = {
        'draw': draw,
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data,
    }
    return JsonResponse(response)

@permission_required('gestion_escolar.acceder_kardex', raise_exception=True)
def kardex_maestro_list(request):
    """Muestra una lista de maestros para seleccionar y ver su kardex."""
    # La lógica ahora está en la vista AJAX, este template solo renderiza la tabla base
    context = {
        'titulo': 'Kardex del Personal'
    }
    return render(request, 'gestion_escolar/kardex_list.html', context)

@permission_required('gestion_escolar.acceder_kardex', raise_exception=True)
def kardex_maestro_detail(request, maestro_id):
    """Muestra una línea de tiempo unificada para un maestro específico."""
    maestro = get_object_or_404(Maestro, pk=maestro_id)
    
    # Obtener el origen para el botón de "volver"
    from_page = request.GET.get('from', 'lista') # 'lista' es el valor por defecto
    
    timeline = []

    # 4. Obtener registros de Historial
    # Incluir registros donde el maestro es el titular o donde aparece como maestro secundario (interino)
    maestro_full_name = f"{maestro.nombres or ''} {maestro.a_paterno or ''} {maestro.a_materno or ''}".strip()
    historial_maestro = Historial.objects.filter(
        Q(maestro=maestro) | Q(maestro_secundario_nombre=maestro_full_name)
    ).select_related('usuario')
    for item in historial_maestro:
        detalle_display = item.motivo or 'Ver documento'
        if item.tipo_documento == "Asignación de Vacancia" and item.datos_tramite:
            vacancia_data = item.datos_tramite.get("detalles_vacancia", {})
            claves_presupuestales = []
            if vacancia_data.get("clave_presupuestal_posicion"):
                claves_presupuestales.append(f"Posición: {vacancia_data['clave_presupuestal_posicion']}")
            if vacancia_data.get("maestro_titular_info", {}).get("clave_presupuestal"):
                claves_presupuestales.append(f"Titular: {vacancia_data['maestro_titular_info']['clave_presupuestal']}")
            if vacancia_data.get("maestro_interino_info", {}).get("clave_presupuestal"):
                claves_presupuestales.append(f"Interino: {vacancia_data['maestro_interino_info']['clave_presupuestal']}")
            
            if claves_presupuestales:
                detalle_display = "Claves Presupuestales: " + ", ".join(claves_presupuestales)
            else:
                detalle_display = "Asignación de Vacancia (sin claves presupuestales registradas)"

        timeline.append({
            'fecha': item.fecha_creacion,
            'tipo': 'Trámite',
            'descripcion': item.tipo_documento,
            'detalle': detalle_display, # Usar el detalle_display modificado
            'usuario': item.usuario.username if item.usuario else 'Sistema',
            'objeto': item
        })

    # 2. Obtener registros de RegistroCorrespondencia
    correspondencia_recibida = RegistroCorrespondencia.objects.filter(maestro=maestro)
    for item in correspondencia_recibida:
        # Convertir date a datetime y hacerlo timezone-aware para unificar el tipo en la timeline
        fecha_dt_naive = datetime.combine(item.fecha_recibido, datetime.min.time())
        fecha_dt_aware = timezone.make_aware(fecha_dt_naive, timezone.get_current_timezone())
        timeline.append({
            'fecha': fecha_dt_aware, # Use the timezone-aware datetime
            'tipo': 'Correspondencia',
            'descripcion': f"Recibido: {item.get_tipo_documento_display()} de {item.remitente}",
            'detalle': item.contenido,
            'usuario': item.quien_recibio or 'N/A',
            'objeto': item
        })

    # 3. Obtener registros de KardexMovimiento
    movimientos_kardex = KardexMovimiento.objects.filter(maestro=maestro).select_related('usuario')
    for item in movimientos_kardex:
        timeline.append({
            'fecha': item.fecha,
            'tipo': 'Movimiento Manual',
            'descripcion': 'Anotación en Kardex',
            'detalle': item.descripcion,
            'usuario': item.usuario.username if item.usuario else 'Sistema',
            'objeto': item
        })

    # Ordenar la línea de tiempo por fecha, de más reciente a más antiguo
    timeline.sort(key=lambda x: x['fecha'], reverse=True)

    context = {
        'maestro': maestro,
        'timeline': timeline,
        'titulo': f'Kardex de {maestro}',
        'from_page': from_page
    }
    
    return render(request, 'gestion_escolar/kardex_detail.html', context)