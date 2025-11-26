"""
Comando para importar FUPs desde un archivo Excel.

Uso:
    python manage.py importar_fups ruta/al/archivo.xlsx

El archivo Excel debe tener las siguientes columnas:
- FECHA
- NOMBRE
- RFC
- CURP
- CLAVE PRESUPUESTAL
- EFECTOS
- FOLIO
- OBSERVACIONES
- SOSTENIMIENTO
- TECHO FINANCIERO
"""

from django.core.management.base import BaseCommand, CommandError
from gestion_escolar.models import FUP, Maestro
import openpyxl
from datetime import datetime
import os


class Command(BaseCommand):
    help = 'Importa FUPs desde un archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument('archivo_excel', type=str, help='Ruta al archivo Excel con los datos de FUPs')

    def handle(self, *args, **options):
        archivo_path = options['archivo_excel']
        
        # Verificar que el archivo existe
        if not os.path.exists(archivo_path):
            raise CommandError(f'El archivo "{archivo_path}" no existe')
        
        self.stdout.write(self.style.SUCCESS(f'Leyendo archivo: {archivo_path}'))
        
        try:
            # Cargar el archivo Excel
            wb = openpyxl.load_workbook(archivo_path)
            ws = wb.active
            
            # Obtener los encabezados (primera fila)
            headers = [cell.value.strip() if cell.value else '' for cell in ws[1]]
            self.stdout.write(f'Columnas encontradas: {headers}')
            
            # Contadores
            creados = 0
            errores = 0
            actualizados = 0
            
            # Procesar cada fila (empezando desde la fila 2)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Crear un diccionario con los datos de la fila
                    datos = {}
                    for header, value in zip(headers, row):
                        if header:
                            datos[header.strip()] = value
                    
                    # Saltar filas vacías
                    if not datos.get('CURP') and not datos.get('RFC'):
                        continue
                    
                    # Buscar el maestro por CURP o RFC
                    maestro = None
                    if datos.get('CURP'):
                        curp = str(datos['CURP']).strip().upper()
                        try:
                            maestro = Maestro.objects.get(curp=curp)
                        except Maestro.DoesNotExist:
                            self.stdout.write(self.style.ERROR(
                                f'Fila {row_num}: Maestro con CURP {curp} no encontrado'
                            ))
                            errores += 1
                            continue
                    elif datos.get('RFC'):
                        rfc = str(datos['RFC']).strip().upper()
                        try:
                            maestro = Maestro.objects.get(rfc=rfc)
                        except Maestro.DoesNotExist:
                            self.stdout.write(self.style.ERROR(
                                f'Fila {row_num}: Maestro con RFC {rfc} no encontrado'
                            ))
                            errores += 1
                            continue
                    else:
                        self.stdout.write(self.style.ERROR(
                            f'Fila {row_num}: No se proporcionó CURP ni RFC'
                        ))
                        errores += 1
                        continue
                    
                    # Procesar la fecha
                    fecha = None
                    if datos.get('FECHA'):
                        if isinstance(datos['FECHA'], datetime):
                            fecha = datos['FECHA'].date()
                        else:
                            # Intentar parsear la fecha si es string
                            try:
                                fecha = datetime.strptime(str(datos['FECHA']), '%d/%m/%Y').date()
                            except ValueError:
                                try:
                                    fecha = datetime.strptime(str(datos['FECHA']), '%Y-%m-%d').date()
                                except ValueError:
                                    try:
                                        # Formato DD-MM-YYYY
                                        fecha = datetime.strptime(str(datos['FECHA']), '%d-%m-%Y').date()
                                    except ValueError:
                                        self.stdout.write(self.style.WARNING(
                                            f'Fila {row_num}: Formato de fecha inválido "{datos["FECHA"]}", usando fecha actual'
                                        ))
                                        fecha = datetime.now().date()
                    else:
                        fecha = datetime.now().date()
                    
                    # Obtener folio
                    folio = str(datos.get('FOLIO', '')).strip() if datos.get('FOLIO') else ''
                    
                    # Verificar si ya existe un FUP con el mismo folio
                    fup_existente = None
                    if folio:
                        try:
                            fup_existente = FUP.objects.get(folio=folio)
                            self.stdout.write(self.style.WARNING(
                                f'Fila {row_num}: Ya existe un FUP con folio {folio}, se actualizará'
                            ))
                        except FUP.DoesNotExist:
                            pass
                    
                    # Preparar datos del FUP
                    # Si el techo financiero está vacío en el Excel, usar el del maestro
                    techo_financiero = str(datos.get('TECHO FINANCIERO', '')).strip() if datos.get('TECHO FINANCIERO') else ''
                    if not techo_financiero and maestro.techo_financiero:
                        techo_financiero = maestro.techo_financiero
                        self.stdout.write(self.style.WARNING(
                            f'Fila {row_num}: Techo financiero vacío, usando el del maestro: {techo_financiero}'
                        ))
                    
                    efectos = str(datos.get('EFECTOS', '')).strip() if datos.get('EFECTOS') else ''
                    sostenimiento = str(datos.get('SOSTENIMIENTO', '')).strip().upper() if datos.get('SOSTENIMIENTO') else ''
                    observaciones = str(datos.get('OBSERVACIONES', '')).strip() if datos.get('OBSERVACIONES') else ''
                    
                    # Validar sostenimiento
                    if sostenimiento and sostenimiento not in ['FEDERAL', 'ESTATAL']:
                        self.stdout.write(self.style.WARNING(
                            f'Fila {row_num}: Sostenimiento "{sostenimiento}" no válido, se dejará vacío'
                        ))
                        sostenimiento = ''
                    
                    if fup_existente:
                        # Actualizar FUP existente
                        fup_existente.maestro = maestro
                        fup_existente.techo_financiero = techo_financiero
                        fup_existente.efectos = efectos
                        fup_existente.sostenimiento = sostenimiento
                        fup_existente.observaciones = observaciones
                        fup_existente.save()
                        
                        if fecha:
                            FUP.objects.filter(pk=fup_existente.pk).update(fecha=fecha)
                        
                        actualizados += 1
                        self.stdout.write(self.style.SUCCESS(
                            f'Fila {row_num}: FUP actualizado - Folio: {folio} - {maestro.nombres} {maestro.a_paterno}'
                        ))
                    else:
                        # Crear nuevo FUP
                        fup = FUP(
                            maestro=maestro,
                            folio=folio,
                            techo_financiero=techo_financiero,
                            efectos=efectos,
                            sostenimiento=sostenimiento,
                            observaciones=observaciones,
                        )
                        
                        fup.save()
                        
                        if fecha:
                            FUP.objects.filter(pk=fup.pk).update(fecha=fecha)
                        
                        creados += 1
                        self.stdout.write(self.style.SUCCESS(
                            f'Fila {row_num}: FUP creado - Folio: {folio} - {maestro.nombres} {maestro.a_paterno}'
                        ))
                    
                except Exception as e:
                    errores += 1
                    self.stdout.write(self.style.ERROR(
                        f'Fila {row_num}: Error al procesar - {str(e)}'
                    ))
            
            # Resumen
            self.stdout.write(self.style.SUCCESS(f'\n=== RESUMEN ==='))
            self.stdout.write(self.style.SUCCESS(f'FUPs creados: {creados}'))
            if actualizados > 0:
                self.stdout.write(self.style.SUCCESS(f'FUPs actualizados: {actualizados}'))
            if errores > 0:
                self.stdout.write(self.style.ERROR(f'Errores: {errores}'))
            
        except Exception as e:
            raise CommandError(f'Error al procesar el archivo: {str(e)}')
